import csv
from functools import wraps
import io
from dotenv import load_dotenv
import os
from functools import lru_cache
import psycopg2
import psycopg2.extras
import re
import json
from datetime import datetime, timedelta
from flask import Flask, request, redirect, jsonify, Response
import requests
import openpyxl
from openpyxl.styles import Font

load_dotenv()

USERNAME = os.getenv("USERNAME")
PASSWORD = os.getenv("PASSWORD")

app = Flask(__name__)

# ── DB helpers ────────────────────────────────────────────────────────────────
DB_URL = os.getenv("DB_URL")  
if DB_URL and DB_URL.startswith("postgres://"):
    DB_URL = DB_URL.replace("postgres://", "postgresql://", 1)

def get_db():
    url = os.getenv("DB_URL")
    if not url:
        print("ERROR: DB_URL environment variable is missing!")
        raise ValueError("DB_URL is not set in the environment")
    conn = psycopg2.connect(url, cursor_factory=psycopg2.extras.DictCursor, sslmode='require')
    return conn

def clean_phone(phone):
    if not phone:
        return None
    digits = re.sub(r'\+', '', phone) 
    if digits.startswith('0'):
        digits = digits[1:]
    elif digits.startswith('44'):
        digits = digits[2:]
    return digits

def get_customer(phone):
    phone_clean = clean_phone(phone)
    if not phone_clean: return None
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM customers WHERE phone = %s LIMIT 1", (phone_clean,))
    row = cur.fetchone()
    cur.close(); conn.close()
    return dict(row) if row else None

def get_last_orders_bulk():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT o.phone, STRING_AGG(oi.quantity || ' x ' || COALESCE(oi.custom_name, p.display_name, p.name), ', ') as summary
        FROM (
            SELECT DISTINCT ON (phone) id, phone
            FROM orders ORDER BY phone, order_date DESC, id DESC
        ) latest
        JOIN orders o ON o.id = latest.id
        JOIN order_items oi ON o.id = oi.order_id
        JOIN products p ON oi.product_code = p.product_code
        GROUP BY o.phone
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return {r["phone"]: r["summary"] for r in rows}

def get_orders(phone, limit=None):
    phone_clean = clean_phone(phone)
    if not phone_clean: return []
    conn = get_db()
    cur = conn.cursor()
    q = '''
        SELECT o.id, o.order_date, o.delivery_date, o.notes, o.is_paid, o.is_dispatched, o.is_delivered,
               p.name, oi.quantity, COALESCE(oi.custom_name, p.display_name, p.name) as product, COALESCE(oi.custom_price, p.price) as price, oi.custom_name, p.product_code
        FROM orders o
        JOIN order_items oi ON o.id = oi.order_id
        JOIN products p ON oi.product_code = p.product_code
        WHERE o.phone = %s
        ORDER BY o.order_date DESC, o.id DESC
    '''
    if limit: q += f" LIMIT {limit}"
    cur.execute(q, (phone_clean,))
    rows = cur.fetchall()
    cur.close(); conn.close()

    orders = {}
    for r in rows:
        oid = r["id"]
        if oid not in orders:
            orders[oid] = {"id": oid, "date": r["order_date"], "delivery_date": r["delivery_date"], "notes": r["notes"], "is_paid": r["is_paid"], "is_dispatched": r["is_dispatched"], "is_delivered": r["is_delivered"], "items": [], "total": 0}
        
        display_name = r["custom_name"] if r["custom_name"] else r["name"]
        orders[oid]["items"].append({"product": display_name, "qty": r["quantity"], "product_code": r["product_code"], "price": r["price"], "custom_name": r["custom_name"]})
        orders[oid]["total"] += (float(r["price"]) or 0) * r["quantity"]

    return list(orders.values())

def get_all_customers():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT c.*, MAX(o.order_date) as last_order_date
        FROM customers c
        LEFT JOIN orders o ON c.phone = o.phone
        GROUP BY c.phone
        ORDER BY last_order_date DESC NULLS LAST, c.name ASC
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]

def get_products_sold(start, end):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT COALESCE(oi.custom_name, p.name) as name, SUM(oi.quantity) as qty, SUM(oi.quantity * COALESCE(oi.custom_price, p.price)) as revenue
        FROM order_items oi
        JOIN orders o ON oi.order_id = o.id
        JOIN products p ON oi.product_code = p.product_code
        WHERE o.order_date BETWEEN %s AND %s
        GROUP BY COALESCE(oi.custom_name, p.name)
        ORDER BY revenue DESC
    """, (start, end))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"name": r["name"], "qty": r["qty"], "revenue": r["revenue"]} for r in rows]

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if request.cookies.get("auth") == "1":
            return f(*args, **kwargs)
        return redirect("/login")
    return wrapper

@lru_cache(maxsize=1)
def get_all_products():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM products WHERE is_active = TRUE ORDER BY sort_order ASC, name ASC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]

def get_delivery_schedule(town):
    if not town: return "No schedule"
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT days FROM delivery_schedules WHERE LOWER(town) = LOWER(%s)", (town,))
    row = cur.fetchone()
    cur.close(); conn.close()
    return row[0] if row else "Unscheduled"

# ── ANALYTICS HELPERS ─────────────────────────
def get_daily_weather_sales(start, end):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT o.order_date, p.gas_type, SUM(oi.quantity) as qty
        FROM orders o
        JOIN order_items oi ON o.id=oi.order_id
        JOIN products p ON oi.product_code=p.product_code
        WHERE o.order_date BETWEEN %s AND %s AND p.gas_type IN ('Butane', 'Propane')
        GROUP BY o.order_date, p.gas_type
    """, (start, end))
    rows = cur.fetchall()
    cur.close(); conn.close()

    weather = {}
    try:
        url = f"https://api.open-meteo.com/v1/forecast?latitude=51.55&longitude=-1.78&start_date={start}&end_date={end}&daily=temperature_2m_mean&timezone=Europe/London"
        r = requests.get(url, timeout=5).json()
        weather = dict(zip(r["daily"]["time"], r["daily"]["temperature_2m_mean"]))
    except Exception as e:
        print(f"Weather API Error: {e}")

    sales_map = {(str(r["order_date"]), r["gas_type"]): int(r["qty"]) for r in rows}
    
    start_dt = datetime.strptime(start, "%Y-%m-%d").date()
    end_dt = datetime.strptime(end, "%Y-%m-%d").date()
    
    data = {"dates": [], "butane": [], "propane": [], "temp": []}
    
    curr = start_dt
    while curr <= end_dt:
        d_str = str(curr)
        data["dates"].append(d_str)
        data["butane"].append(sales_map.get((d_str, "Butane"), 0))
        data["propane"].append(sales_map.get((d_str, "Propane"), 0))
        data["temp"].append(weather.get(d_str, None)) 
        curr += timedelta(days=1)

    return data

def get_period_revenue(start, end):
    conn = get_db()
    cur = conn.cursor()
    
    # 1. Get matched revenue grouped by Order Type for the period
    cur.execute("""
        SELECT COALESCE(o.order_type, 'Delivery') as type, SUM(COALESCE(oi.custom_price, p.price) * oi.quantity)
        FROM orders o JOIN order_items oi ON o.id=oi.order_id JOIN products p ON oi.product_code=p.product_code
        WHERE o.order_date BETWEEN %s AND %s
        GROUP BY o.order_type
    """,(start, end))
    results = {r[0]: float(r[1]) for r in cur.fetchall()}
    
    # 2. Get Raw SumUp Total for the period
    cur.execute("SELECT SUM(amount) FROM sumup_payments WHERE DATE(created_at) BETWEEN %s AND %s", (start, end))
    raw_sumup = cur.fetchone()[0] or 0.0
    
    # 3. Calculate driver expected cash for the period
    cur.execute("""
        SELECT SUM(COALESCE(oi.custom_price, p.price) * oi.quantity)
        FROM orders o 
        JOIN order_items oi ON o.id=oi.order_id 
        JOIN products p ON oi.product_code=p.product_code
        WHERE o.delivery_date BETWEEN %s AND %s 
          AND o.is_paid IS NOT TRUE 
          AND o.order_type = 'Delivery'
    """, (start, end))
    driver_cash = cur.fetchone()[0] or 0.0

    cur.close(); conn.close()
    
    orders = results.get('Delivery', 0)
    walkin = results.get('Walk-in', 0)
    sumup_matched = results.get('SumUp', 0)
    
    return orders, walkin, sumup_matched, float(raw_sumup), float(driver_cash)

def predict_next_calls(days=3):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT phone, order_date FROM orders GROUP BY phone, order_date ORDER BY phone, order_date")
    rows = cur.fetchall()
    cur.close(); conn.close()
    
    from collections import defaultdict
    data = defaultdict(list)
    for r in rows: data[r["phone"]].append(r["order_date"])

    try:
        w_url = "https://api.open-meteo.com/v1/forecast?latitude=51.55&longitude=-1.78&daily=temperature_2m_mean&forecast_days=14&timezone=Europe/London"
        temps = requests.get(w_url, timeout=3).json()["daily"]["temperature_2m_mean"]
        upcoming_avg_temp = sum(temps) / len(temps)
    except:
        upcoming_avg_temp = 10.0 

    predictions = {i: [] for i in range(days)}
    missed_calls = []
    today = datetime.today().date()

    for phone, dates in data.items():
        if len(dates) < 2: continue
        diffs = [(dates[i]-dates[i-1]).days for i in range(1,len(dates))]
        base_avg = sum(diffs)/len(diffs)

        temp_diff = upcoming_avg_temp - 10.0 
        adjustment_factor = 1.0 + (temp_diff * 0.02)
        adjustment_factor = max(0.7, min(1.3, adjustment_factor))
        adjusted_avg = round(base_avg * adjustment_factor)
        next_date = dates[-1] + timedelta(days=max(1, adjusted_avg))

        cust = get_customer(phone)
        name = cust["name"] if cust else "Unknown"
        call_data = {"name": name, "phone": phone, "expected": str(next_date)}

        if next_date < today:
            missed_calls.append(call_data)

        for i in range(days):
            if next_date == today + timedelta(days=i):
                predictions[i].append(call_data)

    return predictions, missed_calls

def get_inventory_status():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT p.name, p.product_code, COALESCE(i.quantity,0) as stock, COALESCE(SUM(oi.quantity),0) as sold_last_week
        FROM products p
        LEFT JOIN inventory i ON p.product_code=i.product_code
        LEFT JOIN order_items oi ON p.product_code=oi.product_code
        LEFT JOIN orders o ON oi.order_id=o.id AND o.order_date >= CURRENT_DATE - INTERVAL '7 days'
        GROUP BY p.name, p.product_code, i.quantity
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    
    result = []
    for r in rows:
        days_left = (r["stock"] / r["sold_last_week"]) * 7 if r["sold_last_week"] > 0 else None
        result.append({"name": r["name"], "code": r["product_code"], "stock": r["stock"], "days_left": round(days_left,1) if days_left else None})
    return result

# ── Shared HTML assets ────────────────────────────────────────────────────────
STYLE = """
<meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Syne:wght@400;600;700;800&display=swap" rel="stylesheet">
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  :root { --bg: #0e0f13; --surface: #16181f; --border: #252830; --accent: #f97316; --accent2: #fb923c; --text: #e8eaf0; --muted: #6b7280; --danger: #ef4444; --success: #22c55e; }
  .light { --bg: #f6f7fb; --surface: #ffffff; --border: #e5e7eb; --text: #111827; --muted: #6b7280; }
  
  body { font-family: 'Syne', sans-serif; background: var(--bg); color: var(--text); min-height: 100vh; font-size: 17px; }
  nav { display: flex; align-items: center; gap: 24px; padding: 16px 32px; background: var(--surface); border-bottom: 1px solid var(--border); position: sticky; top: 0; z-index: 100; flex-wrap: wrap; }
  nav .logo { font-size: 1.3rem; font-weight: 800; color: var(--accent); text-decoration: none; margin-right: auto; }
  nav a { font-size: 1.05rem; font-weight: 600; color: var(--muted); text-decoration: none; padding: 6px 14px; border-radius: 6px; transition: all .15s; }
  nav a:hover { color: var(--text); background: var(--border); }
  
  .page { max-width: 960px; margin: 0 auto; padding: 36px 24px; }
  .page-wide { max-width: 1400px; margin: 0 auto; padding: 36px 24px; }
  
  h1 { font-size: 2.2rem; font-weight: 800; margin-bottom: 16px;}
  h2 { font-size: 1.6rem; font-weight: 700; }
  h3 { font-size: 1.1rem; font-weight: 700; color: var(--muted); text-transform: uppercase; margin-bottom: 12px; }
  
  .card { background: var(--surface); border: 1px solid var(--border); border-radius: 10px; padding: 24px; }
  .customer-hero { background: linear-gradient(135deg, var(--surface), var(--bg)); border: 1px solid var(--border); border-radius: 10px; padding: 28px 32px; display: flex; align-items: center; gap: 20px; }
  .avatar { width: 68px; height: 68px; background: var(--accent); border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 1.8rem; font-weight: 800; color: #fff; flex-shrink: 0; }
  .badge { display: inline-block; background: rgba(249,115,22,.15); color: var(--accent); border: 1px solid rgba(249,115,22,.3); font-family: 'DM Mono', monospace; font-size: 1.2rem; padding: 4px 12px; border-radius: 99px; margin-left: 8px; }
  
  .btn { display: inline-flex; align-items: center; justify-content: center; gap: 6px; font-family: 'Syne', sans-serif; font-size: 1.05rem; font-weight: 700; padding: 12px 24px; border-radius: 8px; border: none; cursor: pointer; text-decoration: none; transition: all 0.2s;}
  .btn-primary { background: var(--accent); color: #fff; }
  .btn-primary:hover { background: var(--accent2); }
  .btn-ghost { background: transparent; color: var(--muted); border: 1px solid var(--border); }
  .btn-ghost:hover { color: var(--text); border-color: var(--text); }
  .btn-danger { background: transparent; color: var(--danger); border: 1px solid var(--danger); padding: 8px 16px;}
  
  .product-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(160px, 1fr)); gap: 14px; margin: 20px 0; }
  .product-tile { border: 2px solid var(--border); border-radius: 10px; padding: 18px 12px; text-align: center; cursor: pointer; position: relative; transition: all 0.2s; }
  .product-tile:hover { transform: translateY(-2px); box-shadow: 0 4px 12px rgba(0,0,0,0.1); }
  .product-tile.selected { border-width: 3px; border-color: var(--accent); font-weight:bold; }
  .p-name { font-size: 1.1rem; font-weight: 600; line-height: 1.3;}
  .p-qty { display: block; margin-top: 10px; font-family: 'DM Mono', monospace; font-size: 2rem; color: var(--accent); min-height: 2.2rem; }
  
  .order-card { background: var(--bg); border: 1px solid var(--border); border-radius: 10px; padding: 18px; margin-bottom: 14px; font-size: 1.05rem;}
  
  .modern-input, input[type=text], input[type=date], input[type=tel], input[type=number] { padding: 12px 16px; border-radius: 8px; border: 1px solid var(--border); background: var(--surface); color: var(--text); font-size: 1.05rem; width: 100%; margin-bottom: 14px; font-family: 'DM Mono', monospace; }
  .modern-input:focus, input:focus { outline: none; border-color: var(--accent); }
  
  table { width: 100%; border-collapse: collapse; font-size: 1.05rem; }
  th { padding: 16px; border-bottom: 2px solid var(--border); text-align: left; color: var(--muted); font-weight: 700;}
  td { padding: 16px; border-bottom: 1px solid var(--border); text-align: left; }
  tr:hover td { background: rgba(255,255,255,0.03); }
  
  .checkbox-lg { transform: scale(1.6); margin-right: 12px; cursor: pointer; }
  
  #toast { position: fixed; bottom: 24px; right: 24px; background: var(--success); color: #fff; font-weight: 700; padding: 12px 22px; border-radius: 8px; font-size: 1rem; opacity: 0; pointer-events: none; transition: opacity .3s; z-index: 999; }
  #toast.show { opacity: 1; }

  @media print {
      nav, .no-print { display: none !important; }
      body, .page-wide { background: white; color: black; margin: 0; padding: 0; font-size: 12pt; }
      .card { border: none; padding: 0; }
      table { width: 100%; border: 1px solid #ccc; }
      th, td { border: 1px solid #ccc; padding: 8px; color: black; }
      tr:nth-child(even) { background-color: #f2f2f2; }
  }
</style>
"""

NAV = """
<nav>
  <a class="logo" href="/">Sleemans</a>
  <a href="/deliveries">Deliveries</a>
  <a href="/schedule">Schedules</a>
  <a href="/analytics">Analytics</a>
  <a href="/cash">POS</a>
  <a href="/inventory">Inventory</a>
  <a href="/reload_cache" title="Reload Data" style="margin-left:auto; font-size:1.2rem;">🔄</a>
  <a href="#" onclick="toggleTheme()" id="theme-icon" style="margin-left:auto">&#9728;</a>
</nav>
"""

TOAST_JS = '''<div id="toast"></div><script>function showToast(m){const t=document.getElementById("toast");t.textContent=m;t.classList.add("show");setTimeout(()=>t.classList.remove("show"),2500);} 
    function toggleTheme(){
        const isLight = document.body.classList.toggle("light");
        localStorage.setItem("theme", isLight ? "light" : "dark");
        document.getElementById("theme-icon").innerHTML = isLight ? "&#9790;" : "&#9728;";
        }
        (function(){
        const saved = localStorage.getItem("theme");
        if(saved==="light"){
            document.body.classList.add("light");
        }
    })();
</script>'''

def page(title, body, wide=False):
    cls = "page-wide" if wide else "page"
    return f'<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">{STYLE}<title>{title} – Sleemans</title></head><body>{NAV}<div class="{cls}">{body}</div>{TOAST_JS}</body></html>'

# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    return redirect("/search")

@app.route("/reload_cache")
@login_required
def reload_cache():
    get_all_products.cache_clear()
    return redirect(request.referrer or "/search")

@app.route("/api/toggle_delivery_status", methods=["POST"])
@login_required
def toggle_delivery_status():
    data = request.json
    order_id = data.get("order_id")
    field = data.get("field") # 'is_dispatched' or 'is_delivered'
    value = data.get("value")
    
    if order_id and field in ['is_dispatched', 'is_delivered']:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(f"UPDATE orders SET {field} = %s WHERE id = %s", (value, order_id))
        conn.commit()
        cur.close(); conn.close()
        return jsonify({"status": "success"})
    return jsonify({"status": "error"}), 400

@app.route("/roll_undelivered", methods=["POST"])
@login_required
def roll_undelivered():
    target_date_str = request.form.get("date")
    if not target_date_str: return redirect("/deliveries")
    
    target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
    
    conn = get_db()
    cur = conn.cursor()
    
    # FIX: Changed <= to = so it ONLY grabs orders specifically for this exact target_date
    cur.execute("""
        SELECT o.id, o.notes, c.town 
        FROM orders o 
        JOIN customers c ON o.phone = c.phone 
        WHERE o.delivery_date = %s AND o.is_delivered = FALSE AND o.order_type = 'Delivery'
    """, (target_date,))
    orders_to_roll = cur.fetchall()
    
    days_map = {"mon": 0, "tue": 1, "wed": 2, "thu": 3, "fri": 4, "sat": 5, "sun": 6}
    
    for o in orders_to_roll:
        sched = get_delivery_schedule(o['town'])
        next_date = target_date + timedelta(days=1) 
        
        if sched and sched != "Unscheduled":
            sched_days = [days_map.get(d.strip().lower()[:3]) for d in sched.split(",") if days_map.get(d.strip().lower()[:3]) is not None]
            if sched_days:
                for i in range(1, 8):
                    check_date = target_date + timedelta(days=i)
                    if check_date.weekday() in sched_days:
                        next_date = check_date
                        break
        
        new_notes = o['notes'] or ""
        new_notes += f" [Rolled from {target_date_str}]"
        
        cur.execute("UPDATE orders SET delivery_date = %s, notes = %s, is_dispatched = FALSE WHERE id = %s", 
                    (next_date, new_notes.strip(), o['id']))

    conn.commit()
    cur.close(); conn.close()
    
    return redirect(f"/deliveries?date={target_date_str}")

@app.route("/delete_customer", methods=["POST"])
@login_required
def delete_customer():
    phone = clean_phone(request.form.get("phone"))
    if phone:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id FROM orders WHERE phone = %s", (phone,))
        order_ids = [row[0] for row in cur.fetchall()]

        if order_ids:
            cur.execute("DELETE FROM order_items WHERE order_id = ANY(%s)", (order_ids,))
            cur.execute("DELETE FROM orders WHERE phone = %s", (phone,))
            
        cur.execute("DELETE FROM customers WHERE phone = %s", (phone,))
        conn.commit()
        cur.close(); conn.close()

    return redirect("/search")

@app.route("/update_delivery_date", methods=["POST"])
@login_required
def update_delivery_date():
    order_id = request.form.get("order_id")
    new_date = request.form.get("new_date")
    return_date = request.form.get("return_date")

    if order_id and new_date:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("UPDATE orders SET delivery_date = %s WHERE id = %s", (new_date, order_id))
        conn.commit()
        cur.close(); conn.close()

    return redirect(f"/deliveries?date={return_date}")

@app.route("/api/optimize_route", methods=["POST"])
@login_required
def optimize_route():
    data = request.json
    postcodes = data.get("postcodes", [])
    depot_postcode = "SN3 4PN"
    
    ORS_API_KEY = os.getenv("ORS_API_KEY")
    if not ORS_API_KEY:
        return jsonify({"error": "ORS_API_KEY is missing from environment variables."}), 400

    try:
        all_pcs = [depot_postcode] + postcodes
        geo_req = requests.post("https://api.postcodes.io/postcodes", json={"postcodes": all_pcs})
        geo_data = geo_req.json().get("result", [])

        coords_map = {}
        for item in geo_data:
            if item and item.get("query") and item.get("result"):
                pc_key = item["query"].replace(" ", "").upper()
                coords_map[pc_key] = [item["result"]["longitude"], item["result"]["latitude"]]

        depot_key = depot_postcode.replace(" ", "").upper()
        if depot_key not in coords_map:
            return jsonify({"error": "Could not locate depot coordinates."}), 400

        jobs = []
        for i, pc in enumerate(postcodes):
            pc_key = pc.replace(" ", "").upper()
            if pc_key in coords_map:
                jobs.append({
                    "id": i + 1,
                    "location": coords_map[pc_key],
                    "description": pc 
                })

        if not jobs:
            return jsonify({"error": "Could not geocode any of the delivery postcodes."}), 400

        payload = {
            "vehicles": [{
                "id": 1,
                "profile": "driving-car",
                "start": coords_map[depot_key],
                "end": coords_map[depot_key]
            }],
            "jobs": jobs
        }

        headers = {
            "Authorization": ORS_API_KEY,
            "Content-Type": "application/json"
        }

        ors_res = requests.post("https://api.openrouteservice.org/optimization", json=payload, headers=headers)
        
        if ors_res.status_code != 200:
            return jsonify({"error": "OpenRouteService failed to calculate.", "details": ors_res.text}), 500

        routes = ors_res.json().get("routes", [])
        if not routes:
            return jsonify({"error": "No optimized route returned."}), 400

        steps = routes[0].get("steps", [])
        optimized_postcodes = []
        
        for step in steps:
            if step.get("type") == "job":
                job_id = step.get("job")
                for j in jobs:
                    if j["id"] == job_id:
                        optimized_postcodes.append(j["description"])
                        break

        return jsonify({"optimized": optimized_postcodes})

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        u = request.form.get("username")
        p = request.form.get("password")
        if u == USERNAME and p == PASSWORD:
            resp = redirect("/search")
            resp.set_cookie("auth", "1", max_age=60*60*24*30)
            return resp
    
    body = '''
    <h1 style="margin-bottom:24px">Login</h1>
    <div class="card">
      <form method="POST">
        <div style="margin-bottom: 18px;">
          <label style="display:block;margin-bottom:8px;font-weight:bold;color:var(--muted)">Username</label>
          <input type="text" name="username" class="modern-input" required>
        </div>
        <div style="margin-bottom: 18px;">
          <label style="display:block;margin-bottom:8px;font-weight:bold;color:var(--muted)">Password</label>
          <input type="password" name="password" class="modern-input" required>
        </div>
        <button class="btn btn-primary" style="width:100%">Login</button>
      </form>
    </div>
    '''
    return page("Login", body)

@app.route("/lookup")
@login_required
def lookup():
    raw = request.args.get("phone","")
    phone = clean_phone(raw)

    if not phone:
        return page("Lookup", f'<div class="card" style="border-color:var(--danger)"><h2>Invalid number</h2><p>{raw}</p></div>')

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("SELECT is_alias_for FROM customers WHERE phone = %s", (phone,))
        alias_row = cur.fetchone()
        if alias_row and alias_row['is_alias_for']:
            cur.close(); conn.close()
            return redirect(f"/lookup?phone={alias_row['is_alias_for']}")
    except Exception:
        conn.rollback() 
    cur.close(); conn.close()

    user = get_customer(phone)
    orders = get_orders(phone) if phone else []
    products = get_all_products()

    if not user:
        return page("Lookup", f'''
        <div class="card" style="border-color:var(--accent)">
            <h2>Unknown caller</h2>
            <p style="font-size:1.2rem;margin:10px 0;">0{phone}</p>
            <div style="display:flex; gap:12px; margin-top: 16px;">
                <a href="/add_customer?phone={phone}" class="btn btn-primary">+ Add New Customer</a>
                <a href="/link_customer?phone={phone}" class="btn btn-ghost" style="border: 1px solid var(--border);">🔗 Link to Existing Account</a>
            </div>
        </div>
        ''')

    initial = (user['name'] or '?')[0].upper()
    sched = get_delivery_schedule(user.get("town"))

    special_prices = {}
    if user.get("has_special_prices"):
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT product_code, price FROM customer_special_prices WHERE phone = %s", (phone,))
        for r in cur.fetchall():
            special_prices[str(r['product_code'])] = float(r['price'])
        cur.close(); conn.close()

    p_map = {}
    d_prices = {}
    for p in products:
        pid = str(p['product_code'])
        p_map[p['name']] = pid
        if p.get('display_name'):
            p_map[p['display_name']] = pid
        d_prices[pid] = float(p['price'])

    order_cards = ""
    for o in orders[:5]:
        tags = "".join(f'<span style="background:var(--surface);border:1px solid var(--border);padding:4px 8px;border-radius:4px;margin-right:6px;display:inline-block;margin-bottom:4px;font-size:0.9rem;">{i["product"]} <strong style="color:var(--accent)">x{i["qty"]}</strong></span>' for i in o["items"])
        paid_str = "<span style='color:var(--success);font-weight:bold;'>Paid</span>" if o['is_paid'] else "<span style='color:var(--danger);font-weight:bold;'>Unpaid</span>"
        
        order_cards += f'''
        <div class="order-card" style="padding:12px; font-size:0.95rem;">
            <div style="display:flex;justify-content:space-between;margin-bottom:8px;">
                <span style="font-family:'DM Mono',monospace;color:var(--muted)">Ord: {o["date"]}</span>
                {paid_str}
            </div>
            <div style="font-family:'DM Mono',monospace;color:var(--muted);margin-bottom:8px;">Del: {o["delivery_date"] or 'N/A'}</div>
            <div>{tags}</div>
            <div style="margin-top:8px;padding-top:8px;border-top:1px solid var(--border);display:flex;justify-content:space-between;align-items:center;">
                <span style="color:var(--muted);font-size:0.85rem;">Notes: {o["notes"] or 'None'}</span>
                <strong style="font-size:1.1rem;">£{o["total"]:.2f}</strong>
            </div>
            <div style="margin-top:8px; display:flex; justify-content:flex-end; gap:8px;">
                <button type="button" class="btn btn-ghost" style="padding:4px 8px;font-size:0.8rem; color:var(--accent);" onclick="editOrder('{o["id"]}')">✏️ Edit</button>
                <form method="POST" action="/delete_order" style="margin:0;">
                    <input type="hidden" name="order_id" value="{o["id"]}">
                    <input type="hidden" name="phone" value="{phone}">
                    <button class="btn btn-ghost" style="padding:4px 8px;font-size:0.8rem">Delete</button>
                </form>
            </div>
        </div>
        '''

    if not order_cards: order_cards = '<div class="card" style="text-align:center;color:var(--muted)">No previous orders.</div>'

    tiles = ""
    for p in products:
        bg_color = p.get('color', 'var(--surface)')
        safe_name = p["name"].replace("'", "\\'") 
        short_name = p.get("display_name") or p["name"] 
        pid_str = str(p['product_code'])
        
        is_special = pid_str in special_prices
        display_price = special_prices[pid_str] if is_special else p["price"]
        price_style = "color:#eab308; font-weight:bold;" if is_special else ""
        
        tiles += f'''
        <div class="product-tile" id="tile-{p["product_code"]}" style="background-color: {bg_color}; padding: 12px 8px; position:relative;" onclick="addQty('{p["product_code"]}', '{safe_name}', {p["price"]})">
            <div class="p-name" style="font-size:1rem;">{short_name}</div>
            <div style="font-size:0.9rem;color:var(--muted);margin-top:4px; display:flex; justify-content:center; align-items:center; gap:6px;">
                £<span id="price-display-{p['product_code']}" style="{price_style}">{display_price}</span>
                <span onclick="event.stopPropagation(); setCustomPrice('{p['product_code']}', {display_price})" style="cursor:pointer; font-size:0.85rem;" title="Edit Special Price">✏️</span>
            </div>
            <div class="p-qty" id="qty-{p["product_code"]}" style="font-size:1.6rem; min-height:1.8rem; margin-top:6px;"></div>
            <span style="position:absolute;top:6px;right:6px;font-size:1.2rem;color:var(--danger);display:none;cursor:pointer;background:var(--bg);border-radius:50%;width:24px;height:24px;line-height:24px;" id="reset-{p["product_code"]}" onclick="event.stopPropagation();resetTile('{p["product_code"]}')">&#x2715;</span>
        </div>
        '''

    body = f'''
    <div style="display:flex;gap:24px;align-items:flex-start;flex-wrap:wrap;">

        <div style="flex: 1 1 35%; min-width: 320px; display:flex; flex-direction:column; gap:24px;">
            <div class="card customer-hero" style="flex-direction:column; align-items:flex-start; gap:16px; padding:24px;">
                <div style="display:flex; align-items:center; gap:16px; width:100%;">
                    <div class="avatar" style="width:56px; height:56px; font-size:1.5rem;">{initial}</div>
                    <div>
                        <h2 style="font-size:1.4rem;">{user["name"]}</h2>
                        <div class="badge" style="margin-left:0; margin-top:4px; font-size:1rem;">0{phone}</div>
                    </div>
                </div>
                <div style="font-size:1rem;color:var(--muted);font-family:'DM Mono',monospace; line-height:1.4;">
                    {user.get("address","")}<br>{user.get("town","")}<br>{user.get("postcode","")}
                </div>
                <div style="color:var(--accent);font-weight:bold;font-size:0.95rem;">🚚 Days: {sched}</div>
                <div style="display:flex; gap:12px; width:100%;">
                    <button class="btn btn-ghost" style="padding:6px 12px;font-size:0.85rem;flex:1;" onclick="getTravelTime('{user.get("postcode","")}')">📍 Calc Time</button> 
                    <a href="/edit_customer?phone={phone}" class="btn btn-ghost" style="padding:6px 12px;font-size:0.85rem;flex:1;text-align:center;">Edit Customer</a>
                </div>
                <button type="button" class="btn btn-ghost" style="width:100%; margin-top:4px; padding:6px 12px; font-size:0.85rem;" onclick="addAliasNumber('{phone}')">🔗 Add Secondary Number</button>
                <form method="POST" action="/delete_customer" style="width:100%; margin-top:4px;" onsubmit="return confirm('WARNING: Are you sure you want to permanently delete this customer? This will also wipe all their historical order data.');">
                    <input type="hidden" name="phone" value="{phone}">
                    <button type="submit" class="btn btn-danger" style="width:100%; padding:6px 12px; font-size:0.85rem;">🗑️ Delete Customer</button>
                </form>
                <span id="travel-time" style="font-weight:bold;font-size:1rem;color:var(--text);width:100%;text-align:center;"></span>
            </div>
            <div>
                <h3 style="margin-bottom:12px;">Last Orders</h3>
                {order_cards}
            </div>
        </div>

        <div style="flex: 1 1 60%; min-width: 400px;" class="card">
            <h3 id="form-title" style="margin-top:0; border-bottom:1px solid var(--border); padding-bottom:12px; margin-bottom:16px;">Select Products</h3>
            <div class="product-grid" style="grid-template-columns: repeat(auto-fill, minmax(110px, 1fr)); gap: 8px; margin-top:0;">
                {tiles}
            </div>
            <form method="POST" action="/save_order" id="order-form" style="margin-top:24px; border-top:1px solid var(--border); padding-top:20px;">
                <input type="hidden" name="phone" value="{phone}">
                <input type="hidden" name="order_id" id="edit-order-id" value="">
                <input type="hidden" name="items" id="items-input">

                <div style="display:flex;gap:12px;margin-bottom:12px;">
                    <div style="flex:1">
                        <label style="display:block;font-size:0.9rem;font-weight:bold;color:var(--muted);margin-bottom:4px;">Order Date</label>
                        <input type="date" name="order_date" class="modern-input" value="{datetime.today().date()}" style="margin:0; padding:8px 12px;">
                    </div>
                    <div style="flex:1">
                        <label style="display:block;font-size:0.9rem;font-weight:bold;color:var(--muted);margin-bottom:4px;">Delivery Date</label>
                        <input type="date" name="delivery_date" class="modern-input" value="{datetime.today().date()}" style="margin:0; padding:8px 12px;">
                    </div>
                </div>

                <div style="display:flex;gap:12px;align-items:flex-end;margin-bottom:20px;">
                    <div style="flex:1">
                        <label style="display:block;font-size:0.9rem;font-weight:bold;color:var(--muted);margin-bottom:4px;">Notes</label>
                        <input type="text" name="notes" class="modern-input" placeholder="Instructions..." style="margin:0; padding:8px 12px;">
                    </div>
                    <div style="background:var(--surface);border:1px solid var(--border);padding:8px 16px;border-radius:8px;height:42px;display:flex;align-items:center;">
                        <label style="display:flex;align-items:center;font-size:1rem;font-weight:bold;cursor:pointer;margin:0;">
                            <input type="checkbox" name="is_paid" class="checkbox-lg" style="transform:scale(1.3); margin-right:8px;"> Paid
                        </label>
                    </div>
                </div>

                <div style="display:flex;justify-content:space-between;align-items:center; background:var(--bg); padding:16px; border-radius:8px; border:1px solid var(--border);">
                    <div style="font-size:1.2rem; font-weight:bold; color:var(--muted);">Total Due:</div>
                    <h2 style="color:var(--success);margin:0;font-size:2.2rem;">£<span id="live-total">0.00</span></h2>
                </div>

                <div style="display:flex; gap:12px; margin-top:16px;">
                    <button type="button" id="cancel-edit-btn" class="btn btn-ghost" style="display:none; padding:16px; font-size:1.2rem;" onclick="cancelEdit()">Cancel Edit</button>
                    <button type="submit" id="save-btn" class="btn btn-primary" style="flex:1; font-size:1.2rem; padding:16px;" onclick="return prepareSubmit()">Save & Confirm Order</button>
                </div>
            </form>
        </div>
    </div>

    <div id="other-modal" style="display:none; position:fixed; top:0; left:0; width:100%; height:100%; background:rgba(0,0,0,0.6); z-index:1000; align-items:center; justify-content:center; backdrop-filter:blur(3px);">
        <div class="card" style="width:100%; max-width:350px; box-shadow: 0 10px 25px rgba(0,0,0,0.5);">
            <h3 style="margin-top:0;">Add Custom Item</h3>
            <label style="display:block; font-size:0.9rem; margin-bottom:4px; color:var(--muted);">Item Name</label>
            <input type="text" id="other-name" class="modern-input" placeholder="e.g. Extra Regulator">
            
            <label style="display:block; font-size:0.9rem; margin-top:12px; margin-bottom:4px; color:var(--muted);">Price (£)</label>
            <input type="number" id="other-price" class="modern-input" placeholder="0.00" step="0.01">
            
            <div style="display:flex; gap:12px; justify-content:flex-end; margin-top:20px;">
                <button type="button" class="btn btn-ghost" onclick="document.getElementById('other-modal').style.display='none'">Cancel</button>
                <button type="button" class="btn btn-primary" onclick="submitOtherModal()">Add to Order</button>
            </div>
        </div>
    </div>

    <script>
    const currentPhone = "{phone}";
    let items = {{}};
    let total = 0.0;
    let pendingOtherId = null;
    
    let customPrices = {json.dumps(special_prices)};
    const productMap = {json.dumps(p_map)};
    const defaultPrices = {json.dumps(d_prices)};
    const recentOrders = {json.dumps({str(o['id']): o for o in orders[:5]}, default=str)};

    function editOrder(orderId) {{
        let o = recentOrders[orderId];
        if(!o) return;

        cancelEdit(); 
        document.getElementById('form-title').innerText = "Editing Order #" + o.id;
        document.getElementById('order-form').action = "/update_order";
        document.getElementById('edit-order-id').value = o.id;

        if (o.items && Array.isArray(o.items)) {{
            o.items.forEach(i => {{
                let id = i.product_code || productMap[i.product]; 
                if (!id) return; 

                let pName = i.custom_name || i.product;
                let pPrice = i.price;
                if (pPrice === undefined || pPrice === null) {{
                    pPrice = customPrices[id] !== undefined ? customPrices[id] : defaultPrices[id];
                }}
                if (isNaN(pPrice)) pPrice = 0;

                items[id] = {{qty: i.qty, price: pPrice, custom_name: pName}};
                total += (pPrice * i.qty);

                let qtyEl = document.getElementById("qty-"+id);
                if(qtyEl) {{
                    qtyEl.textContent = items[id].qty;
                    document.getElementById("tile-"+id).classList.add("selected");
                    document.getElementById("reset-"+id).style.display = "block";
                }}
            }});
        }}

        document.getElementById("live-total").innerText = Math.max(0, total).toFixed(2);
        let oDate = o.date ? o.date.toString().substring(0, 10) : '';
        let dDate = o.delivery_date ? o.delivery_date.toString().substring(0, 10) : '';

        document.querySelector('input[name="order_date"]').value = oDate;
        document.querySelector('input[name="delivery_date"]').value = dDate;
        document.querySelector('input[name="notes"]').value = o.notes || '';
        document.querySelector('input[name="is_paid"]').checked = o.is_paid;

        let saveBtn = document.getElementById('save-btn');
        saveBtn.innerText = "Update Order";
        saveBtn.style.background = "#eab308"; 
        saveBtn.style.color = "black";
        document.getElementById('cancel-edit-btn').style.display = 'inline-block';
        document.getElementById('form-title').scrollIntoView({{behavior: "smooth"}});
    }}

    function cancelEdit() {{
        for (let id in items) {{
            let qtyEl = document.getElementById("qty-"+id);
            if(qtyEl) qtyEl.textContent = "";
            let tileEl = document.getElementById("tile-"+id);
            if(tileEl) tileEl.classList.remove("selected");
            let resetEl = document.getElementById("reset-"+id);
            if(resetEl) resetEl.style.display = "none";
        }}
        
        items = {{}};
        total = 0.0;
        document.getElementById("live-total").innerText = "0.00";

        document.querySelector('input[name="order_date"]').value = "{datetime.today().date()}";
        document.querySelector('input[name="delivery_date"]').value = "{datetime.today().date()}";
        document.querySelector('input[name="notes"]').value = "";
        document.querySelector('input[name="is_paid"]').checked = false;
        document.getElementById('edit-order-id').value = "";

        document.getElementById('form-title').innerText = "Select Products";
        document.getElementById('order-form').action = "/save_order";
        
        let saveBtn = document.getElementById('save-btn');
        saveBtn.innerText = "Save & Confirm Order";
        saveBtn.style.background = "var(--primary)"; 
        saveBtn.style.color = "white";
        document.getElementById('cancel-edit-btn').style.display = 'none';
    }}

    async function setCustomPrice(id, defaultPrice) {{
        let current = customPrices[id] !== undefined ? customPrices[id] : defaultPrice;
        let newPriceStr = prompt("Enter special permanent price (£) for this customer:", current);
        
        if(newPriceStr === null || newPriceStr.trim() === "") return;
        let newP = parseFloat(newPriceStr);
        if(isNaN(newP) || newP < 0) return alert("Invalid price.");
        
        customPrices[id] = newP;
        let disp = document.getElementById("price-display-"+id);
        if(disp) {{
            disp.innerText = newP.toFixed(2);
            disp.style.color = "#eab308"; 
            disp.style.fontWeight = "bold";
        }}

        try {{
            await fetch('/api/set_special_price', {{
                method: 'POST',
                headers: {{'Content-Type': 'application/json'}},
                body: JSON.stringify({{phone: currentPhone, product_id: id, price: newP}})
            }});
        }} catch(e) {{
            console.error("Failed to sync special price to database");
        }}

        if(items[id]) {{
            let oldLineTotal = items[id].price * items[id].qty;
            let newLineTotal = newP * items[id].qty;
            total = total - oldLineTotal + newLineTotal;
            items[id].price = newP;
            document.getElementById("live-total").innerText = Math.max(0, total).toFixed(2);
        }}
    }}

    function addQty(id, name, price) {{
        if(name === 'Other' && !items[id]) {{
            pendingOtherId = id;
            document.getElementById('other-modal').style.display = 'flex';
            document.getElementById('other-name').value = '';
            document.getElementById('other-price').value = '';
            document.getElementById('other-name').focus();
            return;
        }}
        executeAdd(id, name, price);
    }}

    function submitOtherModal() {{
        let n = document.getElementById('other-name').value.trim();
        let p = parseFloat(document.getElementById('other-price').value);
        if(!n || isNaN(p)) {{
            alert("Please enter a valid name and price.");
            return;
        }}
        
        let id = pendingOtherId;
        items[id] = {{qty:0, price: p, custom_name: n}};
        executeAdd(id, 'Other', p);
        document.getElementById('other-modal').style.display = 'none';
    }}

    function executeAdd(id, name, price) {{
        let cPrice;
        if (name === 'Other') {{
            cPrice = items[id] ? items[id].price : price;
        }} else {{
            cPrice = customPrices[id] !== undefined ? customPrices[id] : price;
        }}
        
        let cName = items[id] ? items[id].custom_name : null;
        
        if(!items[id]) items[id] = {{qty:0, price: cPrice, custom_name: cName}};
        items[id].qty += 1;
        total += cPrice;
        
        document.getElementById("qty-"+id).textContent = items[id].qty;
        document.getElementById("tile-"+id).classList.add("selected");
        document.getElementById("reset-"+id).style.display = "block";
        document.getElementById("live-total").innerText = total.toFixed(2);
    }}

    function resetTile(id) {{
        if(items[id]) total -= (items[id].price * items[id].qty);
        delete items[id];
        let qtyEl = document.getElementById("qty-"+id);
        if(qtyEl) qtyEl.textContent="";
        let tileEl = document.getElementById("tile-"+id);
        if(tileEl) tileEl.classList.remove("selected");
        let resetEl = document.getElementById("reset-"+id);
        if(resetEl) resetEl.style.display = "none";
        document.getElementById("live-total").innerText = Math.max(0, total).toFixed(2);
    }}

    function prepareSubmit() {{
        if(!Object.keys(items).length) {{
            alert("Please select at least one product before saving.");
            return false;
        }}
        document.getElementById("items-input").value = JSON.stringify(items);
        return true;
    }}

    async function getTravelTime(postcode) {{
        if(!postcode) return alert("No postcode saved!");
        document.getElementById('travel-time').innerText = "Calc...";
        try {{
            let res = await fetch('/api/travel_time?dest=' + encodeURIComponent(postcode));
            let data = await res.json();
            document.getElementById('travel-time').innerText = "≈ " + (data.time || "Error");
        }} catch(e) {{
            document.getElementById('travel-time').innerText = "Net Error";
        }}
    }}

    function addAliasNumber(masterPhone) {{
        let alias = prompt("Enter the new phone number you want to link to this account:");
        if (alias && alias.trim() !== "") {{
            let form = document.createElement('form');
            form.method = 'POST';
            form.action = '/quick_alias';
            
            let mInput = document.createElement('input');
            mInput.type = 'hidden';
            mInput.name = 'master_phone';
            mInput.value = masterPhone;
            
            let aInput = document.createElement('input');
            aInput.type = 'hidden';
            aInput.name = 'new_alias';
            aInput.value = alias;
            
            form.appendChild(mInput);
            form.appendChild(aInput);
            document.body.appendChild(form);
            form.submit();
        }}
    }}
    </script>
    '''
    return page("Lookup", body, wide=True)

@app.route("/api/set_special_price", methods=["POST"])
@login_required
def set_special_price():
    data = request.json
    phone = data.get("phone")
    pid = data.get("product_id") 
    price = data.get("price")
    
    if phone and pid is not None and price is not None:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute("UPDATE customers SET has_special_prices = TRUE WHERE phone = %s", (phone,))
        cur.execute("""
            INSERT INTO customer_special_prices (phone, product_code, price)
            VALUES (%s, %s, %s)
            ON CONFLICT (phone, product_code) DO UPDATE SET price = EXCLUDED.price
        """, (phone, pid, price))
        
        conn.commit()
        cur.close()
        conn.close()
        
    return jsonify({"status": "success"})

@app.route("/link_customer", methods=["GET", "POST"])
@login_required
def link_customer():
    alias_phone = clean_phone(request.args.get("phone", ""))
    
    if request.method == "POST":
        primary_phone = clean_phone(request.form.get("primary_phone"))
        alias_phone = clean_phone(request.form.get("alias_phone"))

        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO customers (phone, name, is_alias_for)
            VALUES (%s, %s, %s)
            ON CONFLICT (phone) DO UPDATE SET is_alias_for = EXCLUDED.is_alias_for
        """, (alias_phone, f"Alias of 0{primary_phone}", primary_phone))
        
        conn.commit()
        cur.close()
        conn.close()
        return redirect(f"/lookup?phone={primary_phone}")

    customers = get_all_customers()
    opts = "".join(f'<option value="{c["phone"]}">0{c["phone"]} - {c["name"]} ({c["postcode"]})</option>' for c in customers if not c.get('is_alias_for'))

    body = f'''
    <h1>Link Phone Number</h1>
    <div class="card" style="max-width: 600px;">
        <p style="margin-bottom:20px; color:var(--muted); line-height:1.5;">
            Attach <strong>0{alias_phone}</strong> to an existing master account. Future calls from this number will automatically redirect to the master profile.
        </p>
        <form method="POST">
            <input type="hidden" name="alias_phone" value="{alias_phone}">
            <label style="display:block;margin-bottom:8px;font-weight:bold;color:var(--muted)">Select Master Customer Account</label>
            <input type="text" id="search-master" class="modern-input" placeholder="🔍 Type name, phone, or postcode to filter..." style="margin-bottom: 8px;">
            <select name="primary_phone" id="master-select" class="modern-input" required style="height: 50px;">
                <option value="">-- Choose Customer --</option>
                {opts}
            </select>
            <div style="display:flex; justify-content:flex-end; gap:12px; margin-top:24px;">
                <a href="/lookup?phone={alias_phone}" class="btn btn-ghost">Cancel</a>
                <button type="submit" class="btn btn-primary">🔗 Link Number</button>
            </div>
        </form>
    </div>
    
    <script>
        const select = document.getElementById('master-select');
        const allOptions = Array.from(select.options);
        
        document.getElementById('search-master').addEventListener('input', function(e) {{
            const term = e.target.value.toLowerCase();
            select.innerHTML = '';
            allOptions.forEach(opt => {{
                if(opt.value === "" || opt.text.toLowerCase().includes(term)) {{
                    select.appendChild(opt);
                }}
            }});
        }});
    </script>
    '''
    return page("Link Customer", body)

@app.route("/api/travel_time")
@login_required
def travel_time():
    dest = request.args.get("dest")
    if not dest: 
        return jsonify({"time": "No destination"})

    base_postcode = "SN3 4PN"
    headers = {"User-Agent": "SleemansCRM/1.0"}

    try:
        base_url = f"https://nominatim.openstreetmap.org/search?postalcode={base_postcode}&country=UK&format=json"
        r_base = requests.get(base_url, headers=headers).json()
        if not r_base: return jsonify({"time": "Base origin not found"})
        base_coords = f"{r_base[0]['lon']},{r_base[0]['lat']}"

        dest_url = f"https://nominatim.openstreetmap.org/search?postalcode={dest}&country=UK&format=json"
        r_dest = requests.get(dest_url, headers=headers).json()
        if not r_dest: return jsonify({"time": "Dest postcode not found"})
        dest_coords = f"{r_dest[0]['lon']},{r_dest[0]['lat']}"

        osrm_url = f"http://router.project-osrm.org/route/v1/driving/{base_coords};{dest_coords}?overview=false"
        r_route = requests.get(osrm_url).json()

        if r_route.get("code") != "Ok": return jsonify({"time": "Routing failed"})

        duration_seconds = r_route["routes"][0]["duration"]
        mins = int(duration_seconds / 60)
        
        if mins > 59:
            hours = mins // 60
            rem_mins = mins % 60
            time_str = f"{hours} hr {rem_mins} mins"
        else:
            time_str = f"{mins} mins"

        return jsonify({"time": time_str})
    except Exception as e:
        return jsonify({"time": "API Error"})

@app.route("/delete_order", methods=["POST"])
@login_required
def delete_order():
    oid = request.form.get("order_id")
    phone = request.form.get("phone") 

    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("SELECT product_code, quantity FROM order_items WHERE order_id=%s", (oid,))
    items = cur.fetchall()
    
    for item in items:
        if item['product_code']:
            cur.execute("""
                INSERT INTO inventory (product_code, quantity) VALUES (%s, %s)
                ON CONFLICT (product_code) DO UPDATE SET quantity = inventory.quantity + EXCLUDED.quantity
            """, (item['product_code'], item['quantity']))

    cur.execute("DELETE FROM order_items WHERE order_id=%s", (oid,))
    cur.execute("DELETE FROM orders WHERE id=%s", (oid,))
    conn.commit()
    cur.close(); conn.close()

    if phone:
        return redirect(f"/lookup?phone={phone}")
    return redirect("/cash") 

@app.route("/save_order", methods=["POST"])
@login_required
def save_order():
    phone = clean_phone(request.form.get("phone"))
    order_date = request.form.get("order_date") or str(datetime.today().date())
    delivery_date = request.form.get("delivery_date") or order_date
    notes = request.form.get("notes", "")
    is_paid = bool(request.form.get("is_paid"))
    
    try: items = json.loads(request.form.get("items","{}"))
    except Exception: items = {}

    if not items: return redirect(f"/lookup?phone={phone}")

    conn = get_db()
    cur = conn.cursor()

    if not get_customer(phone):
        cur.execute("INSERT INTO customers (phone,name) VALUES (%s,%s)", (phone,"Unknown"))

    cur.execute(
        "INSERT INTO orders (phone, order_date, delivery_date, notes, is_paid, order_type) VALUES (%s,%s,%s,%s,%s,'Delivery') RETURNING id",
        (phone, order_date, delivery_date, notes, is_paid)
    )
    oid = cur.fetchone()[0]

    for pid, data in items.items():
        qty = int(data['qty'])
        prod_code = pid if not str(pid).isdigit() else f"SKU-{pid}" # Handle transition
        cur.execute("INSERT INTO order_items (order_id, product_code, quantity, custom_name, custom_price) VALUES (%s,%s,%s,%s,%s)",
                    (oid, pid, qty, data.get('custom_name'), data.get('price')))
        cur.execute("""
            INSERT INTO inventory (product_code, quantity) VALUES (%s, %s)
            ON CONFLICT (product_code) DO UPDATE SET quantity = inventory.quantity + EXCLUDED.quantity
        """, (pid, -qty))

    conn.commit(); cur.close(); conn.close()
    return redirect(f"/lookup?phone={phone}")

@app.route("/update_order", methods=["POST"])
@login_required
def update_order():
    order_id = request.form.get("order_id")
    phone = request.form.get("phone")
    items_json = request.form.get("items", "{}")
    order_date = request.form.get("order_date")
    
    delivery_date = request.form.get("delivery_date")
    if not delivery_date or delivery_date.strip() == "":
        delivery_date = None
        
    notes = request.form.get("notes", "")
    is_paid = True if request.form.get("is_paid") else False

    try:
        items = json.loads(items_json)
    except json.JSONDecodeError:
        items = {}

    if not order_id or not items:
        return redirect(f"/lookup?phone={phone}")

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE orders 
            SET order_date = %s, delivery_date = %s, notes = %s, is_paid = %s
            WHERE id = %s
        """, (order_date, delivery_date, notes, is_paid, order_id))
        
        cur.execute("DELETE FROM order_items WHERE order_id = %s", (order_id,))
        
        for pid, info in items.items():
            qty = int(info.get("qty", 0))
            price = float(info.get("price", 0.0))
            custom_name = info.get("custom_name")
            
            cur.execute("""
                INSERT INTO order_items (order_id, product_code, quantity, custom_price, custom_name)
                VALUES (%s, %s, %s, %s, %s)
            """, (order_id, pid, qty, price, custom_name))
        
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"CRITICAL ERROR updating order {order_id}: {e}")
    finally:
        cur.close()
        conn.close()

    return redirect(f"/lookup?phone={phone}")

@app.route("/deliveries")
@login_required
def deliveries():
    today = datetime.today().date()
    target_date = request.args.get("date", str(today))
    
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("SELECT product_code as id, display_name, name, gas_type, COALESCE(net_weight, 0) as net, COALESCE(gross_weight, 0) as gross FROM products WHERE gas_type IN ('Butane', 'Propane') ORDER BY sort_order ASC, name ASC")
    gas_products = [dict(r) for r in cur.fetchall()]
    butane = [p for p in gas_products if p['gas_type'] == 'Butane']
    propane = [p for p in gas_products if p['gas_type'] == 'Propane']
    
    cur.execute("""
        SELECT p.product_code as id, SUM(oi.quantity) as qty
        FROM orders o JOIN order_items oi ON o.id = oi.order_id JOIN products p ON oi.product_code = p.product_code
        WHERE o.delivery_date = %s AND p.gas_type IN ('Butane', 'Propane')
        GROUP BY p.product_code
    """, (target_date,))
    delivered_map = {str(r['id']): r['qty'] for r in cur.fetchall()}
    
    cur.execute("""
        SELECT o.id, c.name, c.phone, c.address, c.town, c.postcode, o.notes, o.is_paid, o.is_dispatched, o.is_delivered,
               STRING_AGG(oi.quantity || ' x ' || COALESCE(oi.custom_name, p.display_name, p.name), ', ') as items
        FROM orders o
        JOIN customers c ON o.phone = c.phone
        JOIN order_items oi ON o.id = oi.order_id
        JOIN products p ON oi.product_code = p.product_code
        WHERE o.delivery_date = %s AND o.order_type = 'Delivery'
        GROUP BY o.id, c.name, c.phone, c.address, c.town, c.postcode, o.notes, o.is_paid, o.is_dispatched, o.is_delivered
        ORDER BY c.town, c.name
    """, (target_date,))
    rows = cur.fetchall()
    cur.close(); conn.close()

    unique_postcodes = []
    for r in rows:
        pc = r['postcode'].strip() if r['postcode'] else ""
        if pc and pc not in unique_postcodes:
            unique_postcodes.append(pc)

    tr = "".join(f'''
        <tr class="delivery-row" data-postcode="{r['postcode'].strip() if r['postcode'] else ''}" data-dispatched="{str(r['is_dispatched']).lower()}">
            <td class="no-print" style="text-align:center;">
                <input type="checkbox" class="checkbox-lg" {"checked" if r['is_dispatched'] else ""} onchange="toggleStatus({r['id']}, 'is_dispatched', this.checked)">
            </td>
            <td>
                <a href="/lookup?phone={r['phone']}" style="text-decoration:none; color:inherit; display:block;" title="Go to Customer Profile">
                    <strong style="color:var(--accent);">{r['name']}</strong><br>
                    <span style="font-family:'DM Mono',monospace; color:var(--muted);">0{r['phone']}</span>
                </a>
            </td>
            <td>{r['address']}<br>{r['town']}, {r['postcode']}</td>
            <td style="font-weight:bold; color:var(--text);">{r['items']}</td>
            <td style="text-align:center; font-weight:bold; font-size:12pt;">{"<span style='color:var(--success)'>✓</span>" if r['is_paid'] else ""}</td>
            <td>{r['notes'] or ''}</td>
            <td class="no-print" style="text-align:center;">
                <input type="checkbox" class="checkbox-lg" {"checked" if r['is_delivered'] else ""} onchange="toggleStatus({r['id']}, 'is_delivered', this.checked)">
            </td>
            <td class="no-print">
                <form method="POST" action="/update_delivery_date" style="display:flex; gap:6px; flex-direction:column;">
                    <input type="hidden" name="order_id" value="{r['id']}">
                    <input type="hidden" name="return_date" value="{target_date}">
                    <input type="date" name="new_date" value="{target_date}" class="modern-input" style="margin:0; padding:6px; font-size:0.85rem;">
                    <button class="btn btn-ghost" style="padding:4px 8px; font-size:0.8rem;">Move Date</button>
                </form>
            </td>
        </tr>
    ''' for r in rows)

    matrix_headers_1 = f"<th></th>"
    if butane: matrix_headers_1 += f"<th colspan='{len(butane)}' style='background:#fde68a; color:black; text-align:center;'>Butane (UN 1011)</th>"
    if propane: matrix_headers_1 += f"<th colspan='{len(propane)}' style='background:#fca5a5; color:black; text-align:center;'>Propane (UN 1978)</th>"
    
    matrix_headers_2 = "<td></td>"
    for p in butane + propane:
        matrix_headers_2 += f"<td style='font-weight:bold; text-align:center;'>{p['display_name'] or p['name']}</td>"

    row_in = "<td><strong>In</strong></td>" + "".join("<td></td>" for _ in butane + propane)
    row_delivered = "<td><strong>Delivered</strong></td>"
    row_truck = "<td><strong>Total on Truck</strong></td>"
    row_net = "<td><strong>Net Wt (tons)</strong></td>"
    row_gross = "<td><strong>Gross Wt (tons)</strong></td>"
    
    js_data = {}
    
    for p in butane + propane:
        pid = str(p['id'])
        d_qty = delivered_map.get(pid, 0)
        js_data[pid] = {'net': float(p['net']), 'gross': float(p['gross'])}
        
        row_delivered += f"<td style='text-align:center;'>{d_qty}</td>"
        row_truck += f"<td style='text-align:center;'><input type='number' id='truck_{pid}' value='{d_qty}' class='matrix-input' oninput='calcMatrix()'></td>"
        row_net += f"<td style='text-align:center;' id='net_{pid}'>0.000</td>"
        row_gross += f"<td style='text-align:center;' id='gross_{pid}'>0.000</td>"

    body = f'''
    <style>
    input[type=number]::-webkit-outer-spin-button, input[type=number]::-webkit-inner-spin-button {{ -webkit-appearance: none; margin: 0; }}
    input[type=number] {{ -moz-appearance: textfield; }}

    @media print {{
        @page {{ size: landscape; margin: 10mm; }}
        body {{ background: white !important; color: black !important; }}
        nav, .no-print {{ display: none !important; }}
        .print-only {{ display: block !important; }}
        .print-table-row {{ display: table-row !important; }}
        * {{ font-size: 10pt !important; }}
        h2 {{ font-size: 14pt !important; margin: 0 0 10px 0 !important; }}
        .card {{ border: none !important; padding: 0 !important; background: transparent !important; box-shadow: none !important; margin-bottom: 20px !important; }}
        table {{ border: 1px solid #000 !important; width: 100%; border-collapse: collapse; page-break-inside: auto; }}
        tr {{ page-break-inside: avoid; page-break-after: auto; }}
        th, td {{ border: 1px solid #000 !important; color: black !important; padding: 6px !important; vertical-align: middle; }}
        .matrix-input {{ border: none !important; background: transparent !important; width: 100%; text-align: center; font-weight: bold; padding:0; }}
        .matrix-container {{ page-break-before: always; break-before: page; }}
        .footer-container {{ border: 2px solid #000; padding: 10px; margin-top: 20px; page-break-inside: avoid; }}
        .footer-grid-3 {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 15px; border-bottom: 1px dashed #000; padding-bottom: 10px; margin-bottom: 10px; }}
        .footer-grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-top: 15px; }}
    }}
    .matrix-input {{ width: 60px; padding: 4px; text-align: center; border: 1px solid var(--border); background: var(--bg); color: var(--text); border-radius: 4px; }}
    .print-only, .print-table-row {{ display: none; }}
    .empty-row td {{ height: 50px; }}
    </style>
    
    <div class="no-print" style="margin-bottom: 24px; display:flex; justify-content:space-between; align-items:flex-end;">
        <h1 style="margin-bottom:16px;">Deliveries / Run Sheet</h1>
        <form method="POST" action="/roll_undelivered" onsubmit="return confirm('Roll forward all undispatched/undelivered orders for {target_date} to the next available day?');">
            <input type="hidden" name="date" value="{target_date}">
            <button class="btn btn-warning" style="background:#eab308; color:#000; border:none; margin-bottom:16px;">▶️ Roll Undelivered to Next Day</button>
        </form>
    </div>

    <form id="run-form" action="/export_delivery_excel" method="GET">
        <div class="card no-print" style="display:flex; gap:20px; align-items:flex-end; flex-wrap:wrap; margin-bottom:24px;">
            <div style="flex:1; min-width:140px;">
                <label style="font-weight:bold;color:var(--muted);display:block;margin-bottom:8px;">Date:</label>
                <input type="date" name="date" class="modern-input" value="{target_date}" style="margin:0;" onchange="window.location.href='/deliveries?date='+this.value">
            </div>
            <div style="flex:1; min-width:140px;">
                <label style="font-weight:bold;color:var(--muted);display:block;margin-bottom:8px;">Print Filter:</label>
                <select id="dispatch-filter" class="modern-input" style="margin:0; padding:12px;" onchange="filterDeliveries()">
                    <option value="all">Show All</option>
                    <option value="undispatched">Only Undispatched</option>
                    <option value="dispatched">Only Dispatched</option>
                </select>
            </div>
            <div style="flex:2; min-width:140px;">
                <label style="font-weight:bold;color:var(--muted);display:block;margin-bottom:8px;">Driver Name:</label>
                <select id="driver_input" name="driver" class="modern-input" style="margin:0; padding:12px;">
                    <option value="Craig Batterton">Craig Batterton</option>
                    <option value="Saksham Singh">Saksham Singh</option>
                    <option value="Rajesh Singh">Rajesh Singh</option>
                </select>
            </div>
            <div style="flex:2; min-width:140px;">
                <label style="font-weight:bold;color:var(--muted);display:block;margin-bottom:8px;">Vehicle Reg:</label>
                <select id="vehicle_input" name="vehicle_reg" class="modern-input" style="margin:0; padding:12px;">
                    <option value="DU14 EWG">DU14 EWG</option>
                    <option value="YS63 VPX">YS63 VPX</option>
                </select>
            </div>
            <div style="display:flex; gap:12px; align-items:flex-end;">
                <button type="button" id="route-btn" class="btn btn-primary" onclick="calculateRoute()" style="background:#8b5cf6;height:48px; border:none; min-width:150px;">🗺️ Auto-Sort</button>
                <button type="button" class="btn btn-ghost" onclick="triggerPrint()" style="height:48px;">🖨️ Print</button>
                <button type="submit" class="btn btn-primary" style="background:var(--success);height:48px;">📊 Excel</button>
            </div>
        </div>

        <div class="card" style="padding:0;overflow:hidden; margin-bottom:16px;">
            <table>
                <thead style="background:var(--surface)">
                    <tr class="print-table-row">
                        <td colspan="8" style="border:none !important; border-bottom:2px solid #000 !important; padding-bottom:10px !important;">
                            <h2 style="margin:0 0 5px 0;">Delivery Run Sheet</h2>
                            <div style="font-weight: bold; display:flex; justify-content:space-between;">
                                <span>Date: {target_date} <span style="font-weight:normal;">(Printed: <span id="print_time"></span>)</span></span>
                                <span>Driver: <span id="print_driver"></span></span>
                                <span>Vehicle: <span id="print_vehicle"></span></span>
                            </div>
                        </td>
                    </tr>
                    <tr>
                        <th class="no-print" style="text-align:center;">Dispatched</th>
                        <th>Customer</th>
                        <th>Address</th>
                        <th>Order Items</th>
                        <th>Paid</th>
                        <th>Notes</th>
                        <th class="no-print" style="text-align:center;">Delivered</th>
                        <th class="no-print">Actions</th>
                    </tr>
                </thead>
                <tbody id="deliveries-tbody">
                    {tr or '<tr><td colspan="8" style="text-align:center;padding:20px;">No deliveries.</td></tr>'}
                    <tr class="empty-row"><td class="no-print"></td><td>&nbsp;</td><td></td><td></td><td></td><td></td><td class="no-print"></td><td class="no-print"></td></tr>
                    <tr class="empty-row"><td class="no-print"></td><td>&nbsp;</td><td></td><td></td><td></td><td></td><td class="no-print"></td><td class="no-print"></td></tr>
                </tbody>
            </table>
        </div>
        
        <div class="card matrix-container" style="padding:0;overflow:hidden;overflow-x:auto;">
            <table style="min-width: 800px;">
                <thead>
                    <tr>{matrix_headers_1}</tr>
                    <tr style="background:var(--surface); border-bottom: 2px solid var(--border);">{matrix_headers_2}</tr>
                </thead>
                <tbody>
                    <tr>{row_in}</tr>
                    <tr style="background:var(--surface)">{row_delivered}</tr>
                    <tr>{row_truck}</tr>
                    <tr>{row_net}</tr>
                    <tr>{row_gross}</tr>
                </tbody>
            </table>
            <div style="display:flex; justify-content:space-around; background:var(--surface); padding:12px; font-weight:bold; font-size:1.1rem; border-top:2px solid var(--border);">
                <div>Total Net Weight: <span id="grand_net" style="color:var(--accent);">0.000</span> tons</div>
                <div>Total Gross Weight: <span id="grand_gross" style="color:var(--accent);">0.000</span> tons</div>
            </div>
        </div>
    </form>
    
    <div class="print-only footer-container">
        <div class="footer-grid-3">
            <div>
                <strong>SHIPPING NAME:</strong> PROPANE and/or BUTANE<br>
                <strong>CLASSIFICATION:</strong> CLASS 2.1<br>
                <strong>UN NUMBER/S:</strong> UN 1978 (PROPANE) & UN 1011 (BUTANE)
            </div>
            <div>
                <strong>CONSIGNOR:</strong> FLO GAS Ltd<br>
                <strong>CONSIGNEE:</strong> AS Per delivery address shown above<br>
                <strong>EMERGENCY ACTION CODE:</strong> 2YE
            </div>
            <div style="text-align:center;">
                <strong>SLEEMANS</strong><br>
                154 SWINDON RD<br>
                STRATTON SWINDON WILTS.
            </div>
        </div>
        <div>
            <div style="font-weight:bold; margin-bottom: 10px;">THERE ARE NO TEMPERATURE CONTROL CONSIDERATIONS.</div>
            <div class="footer-grid-2">
                <div style="border-top: 1px solid #000; padding-top: 5px;">FOR OR ON BEHALF OF CONSIGNOR: Signature:</div>
                <div style="border-top: 1px solid #000; padding-top: 5px;">DRIVER SIGNATURE:</div>
            </div>
        </div>
    </div>

    <script>
    const pData = {json.dumps(js_data)};

    async function toggleStatus(orderId, field, value) {{
        try {{
            await fetch('/api/toggle_delivery_status', {{
                method: 'POST',
                headers: {{'Content-Type': 'application/json'}},
                body: JSON.stringify({{order_id: orderId, field: field, value: value}})
            }});
            if (field === 'is_dispatched') {{
                // Update row attr for filtering
                let checkbox = document.querySelector(`input[onchange*="${{orderId}}"][onchange*="is_dispatched"]`);
                if(checkbox) checkbox.closest('tr').setAttribute('data-dispatched', value.toString());
            }}
        }} catch(e) {{
            console.error("Failed to update status");
        }}
    }}

    function filterDeliveries() {{
        let filterVal = document.getElementById('dispatch-filter').value;
        let rows = document.querySelectorAll('.delivery-row');
        rows.forEach(row => {{
            let isDisp = row.getAttribute('data-dispatched') === 'true';
            if (filterVal === 'all') row.style.display = '';
            else if (filterVal === 'dispatched' && isDisp) row.style.display = '';
            else if (filterVal === 'undispatched' && !isDisp) row.style.display = '';
            else row.style.display = 'none';
        }});
    }}

    function calcMatrix() {{
        let totalNet = 0.0;
        let totalGross = 0.0;
        
        for (let pid in pData) {{
            let input = document.getElementById('truck_' + pid);
            if(input) {{
                let qty = parseInt(input.value) || 0;
                let net = qty * pData[pid].net;
                let gross = qty * pData[pid].gross;
                
                document.getElementById('net_' + pid).innerText = net.toFixed(3);
                document.getElementById('gross_' + pid).innerText = gross.toFixed(3);
                
                totalNet += net;
                totalGross += gross;
            }}
        }}
        
        document.getElementById('grand_net').innerText = totalNet.toFixed(3);
        document.getElementById('grand_gross').innerText = totalGross.toFixed(3);
    }}

    calcMatrix();

    function triggerPrint() {{
        const now = new Date();
        document.getElementById('print_time').innerText = now.toLocaleTimeString([], {{hour: '2-digit', minute:'2-digit'}});
        document.querySelectorAll('#print_driver').forEach(d => d.innerText = document.getElementById('driver_input').value);
        document.querySelectorAll('#print_vehicle').forEach(v => v.innerText = document.getElementById('vehicle_input').value);
        window.print();
    }}

    async function calculateRoute() {{
        const postcodes = {json.dumps(unique_postcodes)};
        if (postcodes.length === 0) return alert("No postcodes found.");
        const btn = document.getElementById('route-btn');
        const originalText = btn.innerText;
        btn.innerText = "⏳ Optimizing..."; btn.disabled = true;

        try {{
            let res = await fetch('/api/optimize_route', {{
                method: 'POST', headers: {{'Content-Type': 'application/json'}},
                body: JSON.stringify({{postcodes: postcodes}})
            }});
            let data = await res.json();
            if (data.error) throw new Error(data.error);

            const tbody = document.getElementById('deliveries-tbody');
            const rows = Array.from(tbody.querySelectorAll('.delivery-row'));
            const emptyRows = Array.from(tbody.querySelectorAll('.empty-row'));

            data.optimized.forEach(optPostcode => {{
                rows.filter(row => row.getAttribute('data-postcode') === optPostcode).forEach(row => tbody.appendChild(row));
            }});
            emptyRows.forEach(row => tbody.appendChild(row));

            btn.innerText = "✅ Sorted!";
            setTimeout(() => {{ btn.innerText = originalText; btn.disabled = false; }}, 3000);
        }} catch (e) {{
            alert("Optimization Failed: " + e.message);
            btn.innerText = originalText; btn.disabled = false;
        }}
    }}
    </script>
    '''
    return page(f"Deliveries {target_date}", body, wide=True)

@app.route("/export_delivery_excel")
@login_required
def export_delivery_excel():
    target_date = request.args.get("date", str(datetime.today().date()))
    driver = request.args.get("driver", "Craig Batterton")
    vehicle_reg = request.args.get("vehicle_reg", "DU14 EWG")
    
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("SELECT product_code as id, display_name, name, gas_type, COALESCE(net_weight, 0) as net, COALESCE(gross_weight, 0) as gross FROM products WHERE gas_type IN ('Butane', 'Propane') ORDER BY sort_order ASC, name ASC")
    gas_products = [dict(r) for r in cur.fetchall()]
    butane = [p for p in gas_products if p['gas_type'] == 'Butane']
    propane = [p for p in gas_products if p['gas_type'] == 'Propane']
    
    cur.execute("""
        SELECT o.id, c.name, c.phone, c.address, c.town, c.postcode, o.notes, o.is_paid,
               oi.quantity, COALESCE(oi.custom_name, p.display_name, p.name) as product_name, p.product_code as pid
        FROM orders o
        JOIN customers c ON o.phone = c.phone
        JOIN order_items oi ON o.id = oi.order_id
        JOIN products p ON oi.product_code = p.product_code
        WHERE o.delivery_date = %s AND o.order_type = 'Delivery'
        ORDER BY c.town, c.name
    """, (target_date,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    
    orders_dict = {}
    delivered_map = {}
    
    for r in rows:
        oid = r['id']
        if oid not in orders_dict:
            orders_dict[oid] = {
                "name": r['name'], "phone": f"0{r['phone']}" if r['phone'] else "",
                "address": f"{r['address']}, {r['town']}, {r['postcode']}",
                "notes": r['notes'] or "", "is_paid": "Paid" if r['is_paid'] else "Unpaid",
                "items": []
            }
        
        orders_dict[oid]["items"].append(f"{r['quantity']} x {r['product_name']}")
        
        pid = str(r['pid'])
        delivered_map[pid] = delivered_map.get(pid, 0) + r['quantity']
            
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Run Sheet {target_date}"
    
    ws.append(["Delivery Run Sheet", f"Date: {target_date}", f"Driver: {driver}", f"Vehicle: {vehicle_reg}"])
    for cell in ws[1]: cell.font = Font(bold=True, size=14)
    ws.append([])
    
    headers = ["Customer", "Phone", "Address", "Order Items", "Status", "Notes"]
    ws.append(headers)
    for cell in ws[3]: cell.font = Font(bold=True)
        
    for oid, o in orders_dict.items():
        ws.append([o['name'], o['phone'], o['address'], ", ".join(o['items']), o['is_paid'], o['notes']])
        
    ws.append([])
    ws.append([])
    
    start_row = ws.max_row + 1
    
    h1 = [""]
    if butane: h1.extend(["Butane (UN 1011)"] + [""]*(len(butane)-1))
    if propane: h1.extend(["Propane (UN 1978)"] + [""]*(len(propane)-1))
    ws.append(h1)
    
    h2 = [""] + [p['display_name'] or p['name'] for p in butane + propane]
    ws.append(h2)
    
    ws.append(["Out"])
    ws.append(["In"])
    
    row_deliv = ["Delivered"] + [delivered_map.get(str(p['id']), 0) for p in butane + propane]
    ws.append(row_deliv)
    
    row_truck = ["Total on Truck"]
    row_net = ["Net Wt (tons)"]
    row_gross = ["Gross Wt (tons)"]
    
    total_net = 0.0
    total_gross = 0.0
    
    for p in butane + propane:
        qty_on_truck = int(request.args.get(f"truck_{p['id']}", delivered_map.get(str(p['id']), 0)))
        row_truck.append(qty_on_truck)
        
        net = qty_on_truck * float(p['net'])
        gross = qty_on_truck * float(p['gross'])
        
        row_net.append(round(net, 3))
        row_gross.append(round(gross, 3))
        
        total_net += net
        total_gross += gross
        
    ws.append(row_truck)
    ws.append(row_net)
    ws.append(row_gross)
    
    ws.append([])
    ws.append(["Total Net Wt:", round(total_net, 3)])
    ws.append(["Total Gross Wt:", round(total_gross, 3)])
    
    for r in range(start_row, ws.max_row + 1):
        for cell in ws[r]:
            if cell.column == 1 or r <= start_row + 1:
                cell.font = Font(bold=True)
                
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    
    out = io.BytesIO()
    wb.save(out)
    return Response(out.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment;filename=Deliveries_{target_date}.xlsx"})

@app.route("/schedule", methods=["GET", "POST"])
@login_required
def schedule():
    conn = get_db()
    cur = conn.cursor()
    
    if request.method == "POST":
        town = request.form.get("town").strip()
        days = request.form.get("days").strip()
        if town and days:
            cur.execute("""
                INSERT INTO delivery_schedules (town, days) 
                VALUES (%s,%s) 
                ON CONFLICT (town) DO UPDATE SET days=EXCLUDED.days
            """, (town, days))
            conn.commit()
    
    cur.execute("SELECT * FROM delivery_schedules ORDER BY town")
    rows = cur.fetchall()
    cur.close(); conn.close()

    tr = "".join(f"<tr><td><strong>{r['town']}</strong></td><td>{r['days']}</td></tr>" for r in rows)
    
    body = f'''
    <h1>Delivery Schedules</h1>
    <div class="card" style="margin-bottom:24px;">
        <h3>Add / Update Route</h3>
        <form method="POST" style="display:flex;gap:16px;align-items:flex-end;">
            <div style="flex:1">
                <label style="display:block;margin-bottom:8px;font-weight:bold;color:var(--muted)">Town / Area</label>
                <input type="text" name="town" class="modern-input" placeholder="e.g., Swindon" required style="margin:0">
            </div>
            <div style="flex:2">
                <label style="display:block;margin-bottom:8px;font-weight:bold;color:var(--muted)">Delivery Days</label>
                <input type="text" name="days" class="modern-input" placeholder="e.g., Mon, Wed, Fri" required style="margin:0">
            </div>
            <button class="btn btn-primary" style="height:49px;">Save Route</button>
        </form>
    </div>
    <div class="card" style="padding:0;overflow:hidden">
        <table>
            <thead style="background:var(--surface)">
                <tr><th style="width:40%">Town</th><th>Delivery Days</th></tr>
            </thead>
            <tbody>
                {tr or '<tr><td colspan="2" style="text-align:center;padding:20px;">No schedules configured.</td></tr>'}
            </tbody>
        </table>
    </div>
    '''
    return page("Schedules", body)

@app.route("/search")
@login_required
def search():
    customers = get_all_customers()
    last_orders_map = get_last_orders_bulk()
    products = get_all_products()

    prod_opts = "".join(f'''
        <label style="margin-right:20px; display:inline-flex; align-items:center; cursor:pointer; font-weight:600; padding:6px 0;">
            <input type="checkbox" value="{(p.get('display_name') or p.get('name') or '').lower()}" class="prod-filter checkbox-lg"> {(p.get('display_name') or p.get('title') or '')}
        </label>
    ''' for p in products)

    rows = ""
    for u in customers:
        name       = u.get('name', '') or ''
        phone      = u.get('phone', '') or ''
        town       = u.get('town', '') or ''
        postcode   = u.get('postcode', '') or ''
        initial    = (name or '?')[0].upper()
        last_order = last_orders_map.get(phone, "")

        rows += f'''<tr data-general="{name.lower()} {phone} {(u.get('address') or '').lower()} {postcode.lower()}" data-order="{last_order.lower()}">
          <td>
            <div style="display:flex;align-items:center;gap:12px">
              <div style="width:40px;height:40px;background:var(--accent);border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:800;font-size:1.1rem;color:white;flex-shrink:0">{initial}</div>
              <strong>{name}</strong>
            </div>
          </td>
          <td><span style="font-family:'DM Mono',monospace;font-size:1.2rem;color:var(--accent);font-weight:600;">0{phone}</span></td>
          <td style="color:var(--muted)">{u.get('address','')}, {town}, {postcode}</td>
          <td style="color:var(--text);font-weight:600;">{last_order}</td>
          <td style="text-align:right;">
            <a href="/lookup?phone={phone}" class="btn btn-primary" style="padding:8px 16px;">View</a>
          </td>
        </tr>'''

    body = f'''
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:24px">
        <h1 style="margin:0">Customers <span style="background:var(--surface);border:1px solid var(--border);padding:4px 12px;border-radius:99px;font-size:1.2rem;vertical-align:middle;color:var(--muted)">{len(customers)}</span></h1>
        <a href="/add_customer" class="btn btn-primary">+ Add New Customer</a>
    </div>

    <div class="card" style="margin-bottom:24px;">
        <div style="display:flex; gap:16px; align-items:center;">
            <input type="text" id="search-input" class="modern-input" placeholder="Search by name, phone, address, postcode..." oninput="filterTable()" style="font-size:1.2rem; padding:14px; margin:0; flex:1;">
            
            <div style="position:relative;">
                <button type="button" class="btn btn-ghost" onclick="toggleDropdown(event)" style="height:52px;">
                    Filter by Product ▼
                </button>
                <div id="filter-dropdown" style="display:none; position:absolute; right:0; top:60px; background:var(--surface); border:1px solid var(--border); padding:16px; border-radius:8px; z-index:100; min-width:250px; box-shadow: 0 4px 12px rgba(0,0,0,0.2);">
                    <h4 style="margin-bottom:12px; color:var(--muted);">Require all selected:</h4>
                    <div id="filters" onchange="filterTable()" style="display:flex; flex-direction:column; gap:8px;">
                        {prod_opts}
                    </div>
                </div>
            </div>
        </div>
    </div>

    <div class="card" style="padding:0;overflow:hidden">
      <table>
        <thead style="background:var(--surface)">
            <tr><th>Name</th><th>Phone</th><th>Location</th><th>Last Order</th><th style="text-align:right">Actions</th></tr>
        </thead>
        <tbody id="table-body">{rows}</tbody>
      </table>
      {'<div class="card" style="text-align:center;color:var(--muted);border:none;">No customers found.</div>' if not customers else ''}
    </div>

    <script>
    function filterTable() {{
        const q = document.getElementById("search-input").value.toLowerCase();
        const checked = Array.from(document.querySelectorAll('.prod-filter:checked')).map(cb => cb.value);

        document.querySelectorAll("#table-body tr").forEach(r => {{
            const gen = r.dataset.general || "";
            const ord = r.dataset.order || "";

            const matchSearch = gen.includes(q) || gen.replace(/^0|^44|\\+/g,"").includes(q);
            const matchProd = checked.every(val => ord.includes(val));

            r.style.display = (matchSearch && matchProd) ? "" : "none";
        }});
    }}

    function toggleDropdown(e) {{
        e.stopPropagation();
        const drop = document.getElementById('filter-dropdown');
        drop.style.display = drop.style.display === 'none' ? 'block' : 'none';
    }}

    document.addEventListener('click', function(e) {{
        const drop = document.getElementById('filter-dropdown');
        if (drop && drop.style.display === 'block' && !drop.contains(e.target)) {{
            drop.style.display = 'none';
        }}
    }});
    </script>'''

    return page("Customers", body, wide=True)

@app.route("/add_customer", methods=["GET","POST"])
@login_required
def add_customer():
    phone_prefill = clean_phone(request.args.get("phone",""))

    if request.method == "POST":
        phone = clean_phone(request.form.get("phone"))
        name  = request.form.get("name")
        address = request.form.get("address")
        town = request.form.get("town")
        postcode = request.form.get("postcode")

        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO customers (phone, name, address, town, postcode)
            VALUES (%s,%s,%s,%s,%s)
            ON CONFLICT (phone) DO UPDATE
            SET name=EXCLUDED.name, address=EXCLUDED.address, town=EXCLUDED.town, postcode=EXCLUDED.postcode
        """, (phone,name,address,town,postcode))
        conn.commit()
        cur.close(); conn.close()

        if "save_and_order" in request.form:
            return redirect(f"/lookup?phone={phone}")
        return redirect("/search")

    body = f'''
    <h1>Add Customer</h1>
    <div class="card">
        <form method="POST">
            <div style="margin-bottom:16px">
                <label style="display:block;font-weight:bold;color:var(--muted);margin-bottom:8px;">Phone Number</label>
                <input type="tel" name="phone" value="{phone_prefill}" required class="modern-input">
            </div>
            <div style="margin-bottom:16px">
                <label style="display:block;font-weight:bold;color:var(--muted);margin-bottom:8px;">Full Name</label>
                <input type="text" name="name" class="modern-input">
            </div>
            <div style="margin-bottom:16px">
                <label style="display:block;font-weight:bold;color:var(--muted);margin-bottom:8px;">Street Address</label>
                <input type="text" name="address" class="modern-input">
            </div>
            <div style="margin-bottom:16px">
                <label style="display:block;font-weight:bold;color:var(--muted);margin-bottom:8px;">Town</label>
                <input type="text" name="town" class="modern-input">
            </div>
            <div style="margin-bottom:16px">
                <label style="display:block;font-weight:bold;color:var(--muted);margin-bottom:8px;">Postcode</label>
                <input type="text" name="postcode" class="modern-input">
            </div>

            <div style="display:flex;gap:16px;justify-content:flex-end;margin-top:32px;border-top:1px solid var(--border);padding-top:24px;">
                <button class="btn btn-ghost" name="save_only">Save & Return</button>
                <button class="btn btn-primary" name="save_and_order">Save & Place Order</button>
            </div>
        </form>
    </div>
    '''
    return page("Add Customer", body)

@app.route("/edit_customer", methods=["GET", "POST"])
@login_required
def edit_customer():
    phone = clean_phone(request.args.get("phone"))

    if request.method == "POST":
        if "cancel" in request.form:
            return redirect(f"/lookup?phone={phone}")

        conn = get_db()
        cur = conn.cursor()
        cur.execute('''
            UPDATE customers
            SET name=%s, address=%s, town=%s, postcode=%s
            WHERE phone=%s
        ''', (
            request.form.get("name"), request.form.get("address"),
            request.form.get("town"), request.form.get("postcode"), phone
        ))
        conn.commit()
        cur.close(); conn.close()
        return redirect(f"/lookup?phone={phone}")

    user = get_customer(phone)

    body = f'''
    <h1>Edit Customer</h1>
    <div class="card">
      <form method="POST">
        <div style="margin-bottom:16px">
            <label style="display:block;font-weight:bold;color:var(--muted);margin-bottom:8px;">Phone (Cannot edit)</label>
            <input type="tel" name="phone" value="{phone}" class="modern-input" disabled style="background:var(--bg)">
        </div>
        <div style="margin-bottom:16px">
            <label style="display:block;font-weight:bold;color:var(--muted);margin-bottom:8px;">Full Name</label>
            <input type="text" name="name" value="{user.get('name','')}" class="modern-input">
        </div>
        <div style="margin-bottom:16px">
            <label style="display:block;font-weight:bold;color:var(--muted);margin-bottom:8px;">Street Address</label>
            <input type="text" name="address" value="{user.get('address','')}" class="modern-input">
        </div>
        <div style="margin-bottom:16px">
            <label style="display:block;font-weight:bold;color:var(--muted);margin-bottom:8px;">Town</label>
            <input type="text" name="town" value="{user.get('town','')}" class="modern-input">
        </div>
        <div style="margin-bottom:16px">
            <label style="display:block;font-weight:bold;color:var(--muted);margin-bottom:8px;">Postcode</label>
            <input type="text" name="postcode" value="{user.get('postcode','')}" class="modern-input">
        </div>

        <div style="display:flex;gap:16px;justify-content:flex-end;margin-top:32px;border-top:1px solid var(--border);padding-top:24px;">
            <button class="btn btn-ghost" name="cancel">Cancel</button>
            <button class="btn btn-primary" name="save">Save Changes</button>
        </div>
      </form>
    </div>
    '''
    return page("Edit Customer", body)

@app.route("/analytics")
@login_required
def analytics():
    today_dt = datetime.today().date()
    default_start = (today_dt - timedelta(days=14)).strftime("%Y-%m-%d")
    today_str = today_dt.strftime("%Y-%m-%d")
    
    start = request.args.get("start") or default_start
    end = request.args.get("end") or today_str

    products_sold = get_products_sold(start, end)
    preds, missed = predict_next_calls(3)
    inventory = get_inventory_status()
    weather_data = get_daily_weather_sales(start, end)

    orders_rev, walkin_rev, sumup_matched_rev, raw_sumup_rev, driver_cash = get_period_revenue(start, end)
    total_matched = orders_rev + walkin_rev + sumup_matched_rev
    true_total = orders_rev + walkin_rev + raw_sumup_rev

    prod_labels = [p['name'] for p in products_sold]
    prod_qtys = [int(p['qty']) for p in products_sold]
    period_rev = sum([float(p['revenue']) for p in products_sold])

    inventory_html = ""
    for i in inventory:
        inventory_html += f'''
        <div style="background:var(--bg);padding:16px;border-radius:8px;border:1px solid var(--border)">
            <strong style="font-size:1.1rem;display:block;margin-bottom:8px;">{i["name"]} <span style="font-size:0.8rem; color:var(--muted)">({i["code"]})</span></strong>
            <span style="font-family:'DM Mono',monospace;font-size:1.2rem;color:var(--accent)">{i["stock"]} left</span>
        </div>
        '''

    call_rows = ""
    all_expected = missed + preds.get(0, [])
    for item in all_expected:
        phone, expected, name = item['phone'], item['expected'], item['name']
        is_late = expected < today_str
        date_color = "var(--danger)" if is_late else "var(--accent)"
        
        call_rows += f'''
        <tr class="predictive-row" data-call-id="{phone}_{expected}">
            <td>
                <a href="/lookup?phone={phone}" style="color:var(--text); text-decoration:none; display:flex; flex-direction:column; gap:4px; padding:4px 0;">
                    <strong style="color:var(--accent); font-size:1.05rem; transition:color 0.2s;">{name}</strong>
                    <span style="color:var(--muted); font-family:'DM Mono',monospace; font-size:0.9rem;">0{phone}</span>
                </a>
            </td>
            <td style="text-align:right; vertical-align:middle;">
                <div style="font-size:0.85rem; color:var(--muted); margin-bottom:8px;">
                    Expected: <strong style="color:{date_color}">{expected}</strong>
                </div>
                <button type="button" class="btn btn-ghost" onclick="dismissCall('{phone}', '{expected}')" style="padding:4px 10px; font-size:0.8rem;">
                    ✓ Dismiss
                </button>
            </td>
        </tr>
        '''
        
    if not call_rows: call_rows = '<tr><td colspan="2" style="text-align:center; padding:16px; color:var(--muted)">No expected calls.</td></tr>'

    body = f"""
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>

    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:24px;">
        <h1 style="margin:0">Analytics & Weather</h1>
        <div style="display:flex;gap:12px;">
            <a href="/download_customers" class="btn btn-ghost">📥 Export Customers</a>
            <a href="/download_orders?start={start}&end={end}" class="btn btn-ghost">📥 Export Orders</a>
            <button type="button" class="btn btn-primary" onclick="alert('Matched Revenue for Period: £{period_rev:.2f}')">💰 Show Revenue</button>
        </div>
    </div>

    <form class="card" style="display:flex;gap:16px;align-items:flex-end;margin-bottom:24px;">
        <div style="flex:1">
            <label style="display:block;font-weight:bold;color:var(--muted);margin-bottom:8px;">Start Date</label>
            <input type="date" name="start" class="modern-input" value="{start}" style="margin:0">
        </div>
        <div style="flex:1">
            <label style="display:block;font-weight:bold;color:var(--muted);margin-bottom:8px;">End Date</label>
            <input type="date" name="end" class="modern-input" value="{end}" style="margin:0">
        </div>
        <button class="btn btn-primary" style="height:51px;">Apply Filter</button>
    </form>

    <div style="display:grid;grid-template-columns:1fr 1fr;gap:24px">

        <div class="card" style="border-color: var(--accent); border-width: 2px;">
            <h3 style="margin-bottom:8px; color: var(--accent);">🚚 Driver End-of-Day Collection</h3>
            <div style="color:var(--muted);margin-bottom:16px;font-size:0.95rem;">
                Estimated cash driver should return with today (Excludes pre-paid).
            </div>
            <div style="font-size: 3rem; font-weight: bold; color: var(--text); text-align: center; margin-top: 20px;">
                £{driver_cash:.2f}
            </div>
        </div>
        
        <div class="card" style="border-color: var(--success); border-width: 2px;">
            <h3 style="margin-bottom:8px; color: var(--success);">Today's Revenue Split (£{true_total:.2f})</h3>
            <div style="color:var(--text);margin-bottom:16px;font-size:0.95rem; font-weight:bold;">
                Tap & Pay Total (SumUp): £{raw_sumup_rev:.2f}
            </div>
            <div style="position:relative; height:200px; width:100%; display:flex; justify-content:center;">
                <canvas id="revenue-split"></canvas>
            </div>
        </div>

        <div class="card" style="grid-column:1 / span 2">
            <h3 style="margin-bottom:16px;">Total Bottles Sold (By Type)</h3>
            <div style="position:relative; height:300px; width:100%;">
                <canvas id="bottles-chart"></canvas>
            </div>
        </div>

        <div class="card" style="grid-column:1 / span 2">
            <h3 style="margin-bottom:16px;">🌤️ Weather Correlation (Gas Type vs Temperature)</h3>
            <div style="position:relative; height:400px; width:100%;">
                <canvas id="weather-correlation"></canvas>
            </div>
        </div>

        <div class="card">
            <h3 style="margin-bottom:8px;">☎️ Expected Calls</h3>
            <div style="color:var(--muted);margin-bottom:16px;font-size:0.95rem;">
                Predictions factor in history + 14-day weather forecast.
            </div>
            <div style="overflow-y:auto; max-height:250px; border:1px solid var(--border); border-radius:8px;">
                <table style="width:100%; font-size:0.95rem; margin:0;">
                    <thead style="background:var(--surface)">
                        <tr>
                            <th style="padding:12px; border-bottom:1px solid var(--border);">Customer Profile</th>
                            <th style="text-align:right; padding:12px; border-bottom:1px solid var(--border);">Status</th>
                        </tr>
                    </thead>
                    <tbody id="predictive-tbody">{call_rows}</tbody>
                </table>
            </div>
        </div>

        <div class="card">
            <h3>📦 Inventory Depletion Warnings</h3>
            <div style="display:grid;grid-template-columns:repeat(auto-fill, minmax(200px, 1fr));gap:16px;margin-top:16px;">
                {inventory_html or '<div style="color:var(--muted)">No inventory data available.</div>'}
            </div>
        </div>

    </div>

    <script>
    function dismissCall(phone, expectedDate) {{
        const callId = phone + '_' + expectedDate;
        let dismissed = JSON.parse(localStorage.getItem('dismissed_calls') || '{{}}');
        dismissed[callId] = true;
        localStorage.setItem('dismissed_calls', JSON.stringify(dismissed));
        hideDismissedRows();
    }}

    function hideDismissedRows() {{
        let dismissed = JSON.parse(localStorage.getItem('dismissed_calls') || '{{}}');
        let rows = document.querySelectorAll('.predictive-row');
        let visibleCount = 0;
        
        rows.forEach(row => {{
            let id = row.getAttribute('data-call-id');
            if (dismissed[id]) row.style.display = 'none';
            else {{ row.style.display = ''; visibleCount++; }}
        }});
        
        const tbody = document.getElementById('predictive-tbody');
        let emptyRow = document.getElementById('empty-predictive-msg');
        
        if (visibleCount === 0 && rows.length > 0) {{
            if (!emptyRow) tbody.innerHTML += '<tr id="empty-predictive-msg"><td colspan="2" style="text-align:center; padding:16px; color:var(--muted)">All expected calls handled.</td></tr>';
        }} else if (emptyRow) emptyRow.remove();
    }}
    document.addEventListener("DOMContentLoaded", hideDismissedRows);

    new Chart(document.getElementById('bottles-chart'), {{
        type: 'bar',
        data: {{
            labels: {json.dumps(prod_labels)},
            datasets: [{{
                label: 'Bottles Sold',
                data: {json.dumps(prod_qtys)},
                backgroundColor: '#3b82f6',
            }}]
        }},
        options: {{ responsive: true, maintainAspectRatio: false }}
    }});

    const weatherData = {json.dumps(weather_data)};
    new Chart(document.getElementById('weather-correlation'), {{
        type: 'bar',
        data: {{
            labels: weatherData.dates,
            datasets: [
                {{
                    label: 'Temperature (°C)',
                    data: weatherData.temp,
                    type: 'line',
                    borderColor: '#ef4444',
                    backgroundColor: '#ef4444',
                    borderWidth: 3,
                    tension: 0.3,
                    yAxisID: 'yTemp'
                }},
                {{
                    label: 'Propane Sold',
                    data: weatherData.propane,
                    backgroundColor: '#fca5a5', 
                    yAxisID: 'yQty'
                }},
                {{
                    label: 'Butane Sold',
                    data: weatherData.butane,
                    backgroundColor: '#fde68a', 
                    yAxisID: 'yQty'
                }}
            ]
        }},
        options: {{ 
            responsive: true,
            maintainAspectRatio: false,
            scales: {{
                x: {{ stacked: true }},
                yQty: {{ 
                    stacked: true, 
                    type: 'linear', 
                    position: 'left', 
                    title: {{ display: true, text: 'Total Bottles Sold' }} 
                }},
                yTemp: {{ 
                    type: 'linear', 
                    position: 'right', 
                    title: {{ display: true, text: 'Avg Daily Temp (°C)' }},
                    grid: {{ drawOnChartArea: false }} 
                }}
            }}
        }}
    }});

    new Chart(document.getElementById('revenue-split'), {{
        type: 'doughnut',
        data: {{
            labels: ['Deliveries', 'Walk-ins', 'SumUp (Matched)', 'SumUp (Unmatched)'],
            datasets: [{{
                data: [{orders_rev}, {walkin_rev}, {sumup_matched_rev}, {max(0, raw_sumup_rev - sumup_matched_rev)}],
                backgroundColor: ['#3b82f6', '#f97316', '#a855f7', '#4b5563'],
                borderColor: ['#2563eb', '#ea580c', '#9333ea', '#374151']
            }}]
        }},
        options: {{ responsive:true, maintainAspectRatio: false }}
    }});
    </script>
    """
    return page("Analytics", body, wide=True)

@app.route("/download_customers")
@login_required
def download_customers():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM customers")
    rows = cur.fetchall()
    cur.close(); conn.close()

    if not rows: return "No data"
    
    dict_rows = [dict(r) for r in rows] 
    si = io.StringIO()
    writer = csv.DictWriter(si, fieldnames=list(dict_rows[0].keys()))
    writer.writeheader()
    writer.writerows(dict_rows)
    return Response(si.getvalue(), mimetype="text/csv", headers={"Content-Disposition":"attachment;filename=customers.csv"})

@app.route("/download_orders")
@login_required
def download_orders():
    start = request.args.get("start")
    end = request.args.get("end")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT o.id, o.phone, o.order_date, o.delivery_date, o.notes, o.is_paid,
               COALESCE(oi.custom_name, p.name) as product, oi.quantity, 
               COALESCE(oi.custom_price, p.price) as unit_price
        FROM orders o
        JOIN order_items oi ON o.id = oi.order_id
        JOIN products p ON oi.product_code = p.product_code
        WHERE o.order_date BETWEEN %s AND %s
        ORDER BY o.order_date DESC
    """,(start,end))
    rows = cur.fetchall()
    cur.close(); conn.close()

    if not rows: return "No data for this period"

    dict_rows = [dict(r) for r in rows] 
    si = io.StringIO()
    writer = csv.DictWriter(si, fieldnames=list(dict_rows[0].keys()))
    writer.writeheader()
    writer.writerows(dict_rows)
    return Response(si.getvalue(), mimetype="text/csv", headers={"Content-Disposition":f"attachment;filename=orders_{start}_{end}.csv"})

@app.route("/cash")
@login_required
def cash():
    products = get_all_products()
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT o.id, o.order_type, o.order_date, o.notes,
               COALESCE(STRING_AGG(oi.quantity || ' x ' || COALESCE(oi.custom_name, p.name), ', '), o.notes) as items,
                COALESCE(SUM(oi.quantity * COALESCE(oi.custom_price, p.price)), 0) as total
        FROM orders o
        LEFT JOIN order_items oi ON o.id = oi.order_id
        JOIN products p ON oi.product_code = p.product_code
        WHERE o.order_type IN ('Walk-in', 'SumUp')
        GROUP BY o.id, o.order_type, o.order_date, o.notes
        ORDER BY o.id DESC LIMIT 15
    """)
    recent = cur.fetchall()
    cur.close(); conn.close()

    tiles = ""
    for p in products:
        bg_color = p.get('color', 'var(--surface)')
        safe_name = p["name"].replace("'", "\\'") 
        short_name = p.get("display_name") or p["name"] 
        
        tiles += f'''
        <div class="product-tile" id="tile-{p["product_code"]}" style="background-color: {bg_color}; padding: 12px 8px; position:relative;" onclick="addQty('{p["product_code"]}', '{safe_name}', {p["price"]})">
            <div class="p-name" style="font-size:1rem;">{short_name}</div>
            <div style="font-size:0.9rem;color:var(--muted);margin-top:4px; display:flex; justify-content:center; align-items:center; gap:6px;">
                £<span id="price-display-{p['product_code']}">{p["price"]}</span>
                <span onclick="event.stopPropagation(); setCustomPrice('{p['product_code']}', {p['price']})" style="cursor:pointer; font-size:0.85rem;" title="Give Discount">✏️</span>
            </div>
            <div class="p-qty" id="qty-{p["product_code"]}" style="font-size:1.6rem; min-height:1.8rem; margin-top:6px;"></div>
            <span style="position:absolute;top:6px;right:6px;font-size:1.2rem;color:var(--danger);display:none;cursor:pointer;background:var(--bg);border-radius:50%;width:24px;height:24px;line-height:24px;" id="reset-{p["product_code"]}" onclick="event.stopPropagation();resetTile('{p["product_code"]}')">&#x2715;</span>
        </div>
        '''

    history_rows = "".join(f'''
        <tr>
            <td style="font-family:'DM Mono',monospace">{r['order_date']}</td>
            <td><span class="badge" style="margin:0; background:{'#3b82f620' if r['order_type']=='SumUp' else '#22c55e20'}; color:{'#3b82f6' if r['order_type']=='SumUp' else '#22c55e'}">{r['order_type']}</span></td>
            <td style="font-weight:bold;">{r['items']}</td>
            <td style="font-weight:bold; color:var(--text);">£{r['total']:.2f}</td>
        </tr>
    ''' for r in recent)

    body = f'''
    <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:24px;">
        <h1 style="margin:0">In-Store Point of Sale</h1>
        <a href="/sync_sumup" class="btn btn-primary" style="background:#3b82f6; border:none;">🔄 Sync SumUp Terminals</a>
    </div>

    <div style="display:flex; gap:24px; align-items:flex-start; flex-wrap:wrap;">
        
        <div style="flex: 1 1 45%; min-width:400px;" class="card">
            <h3 style="margin-top:0; border-bottom:1px solid var(--border); padding-bottom:12px; margin-bottom:16px;">Walk-In Cash Order</h3>
            <div class="product-grid" style="grid-template-columns: repeat(auto-fill, minmax(130px, 1fr)); gap: 10px; margin-top:0;">
                {tiles}
            </div>

            <form method="POST" action="/save_walkin" id="order-form" style="margin-top:24px; border-top:1px solid var(--border); padding-top:20px;">
                <input type="hidden" name="items" id="items-input">
                <input type="text" name="notes" class="modern-input" placeholder="Optional Notes..." style="margin-bottom:16px;">

                <div style="display:flex;justify-content:space-between;align-items:center; background:var(--bg); padding:16px; border-radius:8px; border:1px solid var(--border);">
                    <div style="font-size:1.2rem; font-weight:bold; color:var(--muted);">Total Due:</div>
                    <h2 style="color:var(--success);margin:0;font-size:2.2rem;">£<span id="live-total">0.00</span></h2>
                </div>
                <button class="btn btn-success" style="width:100%;font-size:1.3rem;padding:16px;margin-top:16px; background:var(--success);" onclick="return prepareSubmit()">💳 Complete Sale</button>
            </form>
        </div>

        <div style="flex: 1 1 45%; min-width:400px;" class="card">
            <h3 style="margin-top:0; border-bottom:1px solid var(--border); padding-bottom:12px; margin-bottom:16px;">Recent POS & SumUp Activity</h3>
            <table style="font-size:0.95rem;">
                <thead style="background:var(--surface)"><tr><th>Date</th><th>Type</th><th>Items</th><th>Total</th></tr></thead>
                <tbody>{history_rows or '<tr><td colspan="4" style="text-align:center;">No recent in-store orders.</td></tr>'}</tbody>
            </table>
        </div>
    </div>

    <script>
    let items = {{}}; 
    let total = 0.0;
    let customPrices = {{}}; 

    function setCustomPrice(id, defaultPrice) {{
        let current = customPrices[id] !== undefined ? customPrices[id] : defaultPrice;
        let newPriceStr = prompt("Enter new custom price (£) for this order:", current);
        
        if(newPriceStr === null || newPriceStr.trim() === "") return;
        let newP = parseFloat(newPriceStr);
        if(isNaN(newP) || newP < 0) return alert("Invalid price.");
        
        customPrices[id] = newP;
        let disp = document.getElementById("price-display-"+id);
        if(disp) {{
            disp.innerText = newP.toFixed(2);
            disp.style.color = "#eab308"; 
            disp.style.fontWeight = "bold";
        }}

        if(items[id]) {{
            let oldLineTotal = items[id].price * items[id].qty;
            let newLineTotal = newP * items[id].qty;
            total = total - oldLineTotal + newLineTotal;
            items[id].price = newP;
            document.getElementById("live-total").innerText = Math.max(0, total).toFixed(2);
        }}
    }}

    function addQty(id, name, price) {{
        let cName = null; 
        let cPrice = customPrices[id] !== undefined ? customPrices[id] : price; 
        
        if(name === 'Other') {{
            cName = prompt("Enter specific product name:"); if(!cName) return;
            let priceStr = prompt("Enter price (£):"); 
            cPrice = parseFloat(priceStr); 
            if(isNaN(cPrice)) return;
            customPrices[id] = cPrice; 
        }}
        
        if(!items[id]) items[id] = {{qty:0, price: cPrice, custom_name: cName}};
        items[id].qty += 1; 
        total += cPrice;
        
        document.getElementById("qty-"+id).textContent = items[id].qty;
        document.getElementById("tile-"+id).classList.add("selected");
        document.getElementById("reset-"+id).style.display = "block";
        document.getElementById("live-total").innerText = total.toFixed(2);
    }}

    function resetTile(id) {{
        if(items[id]) total -= (items[id].price * items[id].qty);
        delete items[id];
        document.getElementById("qty-"+id).textContent="";
        document.getElementById("tile-"+id).classList.remove("selected");
        document.getElementById("reset-"+id).style.display = "none";
        document.getElementById("live-total").innerText = Math.max(0, total).toFixed(2);
    }}

    function prepareSubmit() {{
        if(!Object.keys(items).length) {{ alert("Select a product."); return false; }}
        document.getElementById("items-input").value = JSON.stringify(items);
        return true;
    }}
    </script>
    '''
    return page("POS", body, wide=True)

@app.route("/inventory", methods=["GET","POST"])
@login_required
def inventory():
    conn = get_db()
    cur = conn.cursor()

    if request.method == "POST":
        for k, v in request.form.items():
            if k.startswith("qty_"):
                pid = k.split("_")[1]
                cur.execute("INSERT INTO inventory (product_code, quantity) VALUES (%s, %s) ON CONFLICT (product_code) DO UPDATE SET quantity=%s", (pid, v, v))
            elif k.startswith("price_"):
                pid = k.split("_")[1]
                cur.execute("UPDATE products SET price=%s WHERE product_code=%s", (v, pid))
            elif k.startswith("color_"):
                pid = k.split("_")[1]
                cur.execute("UPDATE products SET color=%s WHERE product_code=%s", (v, pid))
            elif k.startswith("net_"):
                pid = k.split("_")[1]
                cur.execute("UPDATE products SET net_weight=%s WHERE product_code=%s", (v or 0, pid))
            elif k.startswith("gross_"):
                pid = k.split("_")[1]
                cur.execute("UPDATE products SET gross_weight=%s WHERE product_code=%s", (v or 0, pid))
            elif k.startswith("gas_"):
                pid = k.split("_")[1]
                cur.execute("UPDATE products SET gas_type=%s WHERE product_code=%s", (v if v in ['Butane', 'Propane'] else None, pid))
            elif k.startswith("display_"):
                pid = k.split("_")[1]
                cur.execute("UPDATE products SET display_name=%s WHERE product_code=%s", (v or None, pid))
            elif k.startswith("sort_"):
                pid = k.split("_")[1]
                cur.execute("UPDATE products SET sort_order=%s WHERE product_code=%s", (v or 0, pid))

        if "new_code" in request.form and request.form["new_code"].strip():
            code = request.form["new_code"].strip()
            name = request.form["new_name"].strip()
            display = request.form.get("new_display") or None
            price = request.form.get("new_price") or 0
            qty = request.form.get("new_qty") or 0
            color = request.form.get("new_color") or 'var(--surface)'
            net = request.form.get("new_net") or 0
            gross = request.form.get("new_gross") or 0
            gtype = request.form.get("new_gas")
            gtype = gtype if gtype in ['Butane', 'Propane'] else None
            sort_order = request.form.get("new_sort") or 0
            
            cur.execute("INSERT INTO products (product_code, name, display_name, price, color, net_weight, gross_weight, gas_type, sort_order) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING", (code, name, display, price, color, net, gross, gtype, sort_order))
            cur.execute("INSERT INTO inventory (product_code, quantity) VALUES (%s,%s) ON CONFLICT DO NOTHING", (code, qty))

        if "delete_pid" in request.form:
            pid = request.form["delete_pid"]
            try:
                cur.execute("DELETE FROM inventory WHERE product_code=%s", (pid,))
                cur.execute("DELETE FROM products WHERE product_code=%s", (pid,))
                conn.commit() 
            except Exception:
                conn.rollback() 
                cur.execute("UPDATE products SET is_active = FALSE WHERE product_code=%s", (pid,))
                cur.execute("DELETE FROM inventory WHERE product_code=%s", (pid,)) 

        conn.commit()
        if hasattr(get_all_products, 'cache_clear'): get_all_products.cache_clear()

    cur.execute("""
        SELECT p.product_code, p.name, p.display_name, p.price, p.color, COALESCE(p.net_weight,0) as net, COALESCE(p.gross_weight,0) as gross, p.gas_type, COALESCE(p.sort_order,0) as sort_order, COALESCE(i.quantity,0) as qty
        FROM products p
        LEFT JOIN inventory i ON p.product_code=i.product_code
        WHERE p.is_active = TRUE
        ORDER BY p.sort_order ASC, p.name ASC
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()

    table_rows = "".join(f'''
        <tr>
            <td><input type="number" name="sort_{r['product_code']}" value="{r['sort_order']}" class="modern-input" style="background:var(--bg);margin:0;width:60px;padding:8px;" disabled></td>
            <td style="font-weight:bold; font-family:'DM Mono',monospace;">{r['product_code']}</td>
            <td><input type="text" name="name_{r['product_code']}" value="{r['name']}" class="modern-input" style="background:var(--bg);margin:0;padding:8px;" disabled></td>
            <td><input type="text" name="display_{r['product_code']}" value="{r['display_name'] or ''}" placeholder="Short" class="modern-input" style="background:var(--bg);margin:0;padding:8px;" disabled></td>
            <td>
                <select name="gas_{r['product_code']}" class="modern-input" style="background:var(--bg);margin:0;padding:8px;" disabled>
                    <option value="" {"selected" if not r['gas_type'] else ""}>Other</option>
                    <option value="Butane" {"selected" if r['gas_type'] == 'Butane' else ""}>Butane</option>
                    <option value="Propane" {"selected" if r['gas_type'] == 'Propane' else ""}>Propane</option>
                </select>
            </td>
            <td><input type="number" name="price_{r['product_code']}" value="{r['price']}" step="0.01" class="modern-input" style="background:var(--bg);margin:0;width:75px;padding:8px;" disabled></td>
            <td><input type="number" name="qty_{r['product_code']}" value="{r['qty']}" class="modern-input" style="background:var(--bg);margin:0;width:75px;padding:8px;" disabled></td>
            <td><input type="number" name="net_{r['product_code']}" value="{r['net']}" step="0.001" class="modern-input" style="background:var(--bg);margin:0;width:80px;padding:8px;" disabled></td>
            <td><input type="number" name="gross_{r['product_code']}" value="{r['gross']}" step="0.001" class="modern-input" style="background:var(--bg);margin:0;width:80px;padding:8px;" disabled></td>
            <td><input type="text" name="color_{r['product_code']}" value="{r['color'] or 'var(--surface)'}" class="modern-input" style="background:var(--bg);margin:0;width:100px;padding:8px;" disabled></td>
            <td style="text-align:right"><button type="submit" name="delete_pid" value="{r['product_code']}" class="btn btn-danger" style="padding:6px 12px;">Delete</button></td>
        </tr>
    ''' for r in rows)

    body = f'''
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:24px;">
        <h1 style="margin:0">Inventory & Products</h1>
        <button class="btn btn-ghost" onclick="enableEdit();return false;">✏️ Edit Current</button>
    </div>
    <div class="card" style="padding:0;overflow:hidden;overflow-x:auto; page-break-before: always;">
        <form method="POST" style="margin:0">
            <table style="min-width: 1100px; font-size:0.95rem;">
                <thead style="background:var(--surface)">
                    <tr><th>Sort</th><th>Code / SKU</th><th>Full Name</th><th>Display Pad</th><th>Gas Type</th><th>Price (£)</th><th>Qty</th><th>Net Wt</th><th>Gross Wt</th><th>Color</th><th style="text-align:right">Actions</th></tr>
                </thead>
                <tbody>{table_rows}</tbody>
            </table>
            <div style="padding:24px;background:var(--surface);border-top:1px solid var(--border)">
                <h3 style="margin-bottom:16px;">Add New Product</h3>
                <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap;">
                    <input type="text" name="new_code" placeholder="Code/SKU" class="modern-input" style="margin:0;width:100px;" required>
                    <input type="text" name="new_name" placeholder="Full Name" class="modern-input" style="margin:0;flex:2;" required>
                    <input type="text" name="new_display" placeholder="Short Name" class="modern-input" style="margin:0;flex:1;">
                    <select name="new_gas" class="modern-input" style="margin:0;flex:1;padding:12px;">
                        <option value="">No Gas Type</option>
                        <option value="Butane">Butane</option>
                        <option value="Propane">Propane</option>
                    </select>
                    <input type="number" name="new_price" placeholder="£" step="0.01" class="modern-input" style="margin:0;width:70px;">
                    <input type="number" name="new_qty" placeholder="Qty" class="modern-input" style="margin:0;width:70px;">
                    <input type="number" name="new_sort" placeholder="Sort" class="modern-input" style="margin:0;width:60px;">
                    <button type="submit" class="btn btn-primary" style="height:49px;">Add</button>
                </div>
            </div>
            <div style="padding:24px;text-align:right;border-top:1px solid var(--border);background:var(--bg)">
                <button type="submit" class="btn btn-success" style="background:var(--success);color:white;padding:16px 32px">Save All Changes</button>
            </div>
        </form>
    </div>
    <script>
    function enableEdit(){{
      document.querySelectorAll('input.modern-input:disabled, select.modern-input:disabled').forEach(i => {{
          i.disabled = false; i.style.background = 'var(--surface)'; i.style.borderColor = 'var(--accent)';
      }});
    }} 
    </script>
    '''
    return page("Inventory", body, wide=True)

@app.route("/sync_sumup")
@login_required
def sync_sumup():
    SUMUP_API_KEY = os.getenv("SUMUP_API_KEY") 
    if not SUMUP_API_KEY: return "API Key missing", 400

    headers = {"Authorization": f"Bearer {SUMUP_API_KEY}"}
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS order_type VARCHAR(50) DEFAULT 'Delivery'")
        cur.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS external_id VARCHAR(255) UNIQUE")
        cur.execute("INSERT INTO customers (phone, name, address, town, postcode) VALUES ('00000000000', 'Walk-in Customer', 'In-Store', 'Swindon', 'SN3 4PN') ON CONFLICT (phone) DO NOTHING")
        cur.execute("INSERT INTO customers (phone, name, address, town, postcode) VALUES ('11111111111', 'SumUp Auto-Sync', 'Digital', 'Swindon', 'SN3 4PN') ON CONFLICT (phone) DO NOTHING")
        conn.commit()

        db_products = {p['name'].strip().lower(): p for p in get_all_products()}
        db_display_names = {p['display_name'].strip().lower(): p for p in get_all_products() if p.get('display_name')}
        
        url = "https://api.sumup.com/v0.1/me/transactions/history?limit=30"
        r = requests.get(url, headers=headers)
        if r.status_code != 200: return f"API Error: {r.text}", 400
            
        items = r.json().get("items", [])
        debug_logs = []
        
        for item in items:
            if item.get("status") == "SUCCESSFUL":
                t_id = item.get("id")
                amount = item.get("amount")
                created_at = item.get("timestamp")
                
                cur.execute("INSERT INTO sumup_payments (id, amount, status, description, created_at) VALUES (%s, %s, %s, %s, %s) ON CONFLICT (id) DO NOTHING", 
                            (t_id, amount, 'PAID', f"SumUp: {item.get('transaction_code')}", created_at))
                
                cur.execute("SELECT id FROM orders WHERE external_id = %s", (t_id,))
                if cur.fetchone(): 
                    debug_logs.append(f"[{t_id}] Skipped: Already synced to CRM.")
                    continue
                
                tx_url = f"https://api.sumup.com/v0.1/transactions/{t_id}"
                res = requests.get(tx_url, headers=headers)

                if res.status_code != 200:
                    debug_logs.append(f"[{t_id}] ❌ Failed fetch ({res.status_code})")
                    continue

                tx_res = res.json()

                cur.execute("SELECT id FROM orders WHERE external_id = %s", (t_id,))
                if cur.fetchone(): 
                    debug_logs.append(f"[{t_id}] Skipped: Already synced to CRM.")
                    continue
                
                products_sold = (
                    tx_res.get("products")
                    or tx_res.get("line_items")
                    or tx_res.get("items")
                    or tx_res.get("cart", {}).get("items")
                    or []
                )
                items_to_insert = []
                
                if not products_sold:
                    order_date = item.get("timestamp", str(datetime.today().date()))[:10]

                    cur.execute(
                        "INSERT INTO orders (phone, order_date, delivery_date, notes, is_paid, order_type, external_id) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
                        ('11111111111', order_date, order_date, f"SumUp Quick Sale: £{amount}", True, 'SumUp', t_id)
                    )
                    oid = cur.fetchone()[0]

                    debug_logs.append(f"[{t_id}] 💰 Created Quick Sale Order #{oid} (no item breakdown)")
                    continue
                    
                for sp in products_sold:
                    sp_name = (sp.get('name') or sp.get('title') or '').strip().lower()
                    matched_product = db_products.get(sp_name) or db_display_names.get(sp_name)
                    if matched_product:
                        items_to_insert.append({
                            'pid': matched_product['product_code'],
                            'qty': sp.get('quantity', 1),
                            'price': sp.get('price', matched_product['price'])
                        })
                        debug_logs.append(f"[{t_id}] ✅ Matched SumUp '{sp_name}' to Database item '{matched_product['name']}'")
                    else:
                        debug_logs.append(f"[{t_id}] ❌ FAILED to match SumUp name '{sp_name}'. Please ensure this EXACT name exists in your CRM Inventory.")
                
                if items_to_insert:
                    order_date = tx_res.get("timestamp", str(datetime.today().date()))[:10]
                    cur.execute(
                        "INSERT INTO orders (phone, order_date, delivery_date, notes, is_paid, order_type, external_id) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
                        ('11111111111', order_date, order_date, f"SumUp: {item.get('transaction_code')}", True, 'SumUp', t_id)
                    )
                    oid = cur.fetchone()[0]
                    for i in items_to_insert:
                        cur.execute("INSERT INTO order_items (order_id, product_code, quantity, custom_price) VALUES (%s,%s,%s,%s)", (oid, i['pid'], i['qty'], i['price']))
                        cur.execute("INSERT INTO inventory (product_code, quantity) VALUES (%s, %s) ON CONFLICT (product_code) DO UPDATE SET quantity = inventory.quantity + EXCLUDED.quantity", 
                                    (i['pid'], -i['qty']))
                    debug_logs.append(f"[{t_id}] 📦 Successfully created Order #{oid} and deducted inventory.")
                else:
                    debug_logs.append(f"[{t_id}] ⚠️ Skipped creating order because no products could be matched.")
            
        conn.commit()
        cur.close(); conn.close()
        
        log_html = "<br><br>".join(debug_logs)
        return f"""
        <body style="font-family: monospace; padding: 40px; background: #0e0f13; color: #e8eaf0; font-size:16px;">
            <h1 style="color: #22c55e;">Sync Diagnostic Complete</h1>
            <p style="margin-bottom: 24px;">Review the logs below to see how SumUp data was processed:</p>
            <div style="background: #16181f; padding: 20px; border: 1px solid #252830; border-radius: 8px; line-height: 1.6;">
                {log_html or "No new SumUp transactions found to process."}
            </div>
            <br>
            <a href="/cash" style="color: #3b82f6; font-size: 1.2rem; font-weight: bold; text-decoration: none;">← Return to POS</a>
        </body>
        """
        
    except Exception as e:
        return f"<body style='padding:40px; background:#0e0f13; color:#ef4444; font-family:monospace;'><h1>CRITICAL DATABASE ERROR</h1><p>{str(e)}</p></body>", 500

@app.route("/save_walkin", methods=["POST"])
@login_required
def save_walkin():
    notes = request.form.get("notes", "")
    try: items = json.loads(request.form.get("items","{}"))
    except: items = {}
    if not items: return redirect("/cash")

    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("INSERT INTO customers (phone, name, address, town, postcode) VALUES ('00000000000', 'Walk-in Customer', 'In-Store', 'Swindon', 'SN3 4PN') ON CONFLICT (phone) DO NOTHING")
    
    cur.execute(
        "INSERT INTO orders (phone, order_date, delivery_date, notes, is_paid, order_type) VALUES (%s, CURRENT_DATE, CURRENT_DATE, %s, True, 'Walk-in') RETURNING id",
        ('00000000000', notes)
    )
    oid = cur.fetchone()[0]

    for pid, data in items.items():
        qty = int(data['qty'])
        prod_code = pid if not str(pid).isdigit() else f"SKU-{pid}"
        cur.execute("INSERT INTO order_items (order_id, product_code, quantity, custom_name, custom_price) VALUES (%s,%s,%s,%s,%s)",
                    (oid, prod_code, qty, data.get('custom_name'), data.get('price')))
        cur.execute("""
            INSERT INTO inventory (product_code, quantity) VALUES (%s, %s)
            ON CONFLICT (product_code) DO UPDATE SET quantity = inventory.quantity + EXCLUDED.quantity
        """, (prod_code, -qty))

    conn.commit(); cur.close(); conn.close()
    return redirect("/cash")
    
@app.route("/api/orders")
@login_required
def api_orders():
    phone = clean_phone(request.args.get("phone", ""))
    return jsonify(get_orders(phone) if phone else [])

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)